"""
Commande pour importer les dépenses depuis la feuille DEPENSES du fichier Excel DATADAF.xlsx
Structure : MOIS, ANNEE, DATE, ARTICLE LITTERA, LIBELLE DEPENSES, BANQUE, MONTANT EN Fc, MONTANT EN $us, OBSERVATION
"""
import os
from decimal import Decimal
from datetime import datetime

from django.core.management.base import BaseCommand
from django.conf import settings

from demandes.models import DepenseFeuille, NatureEconomique
from banques.models import Banque


class Command(BaseCommand):
    help = 'Importe les dépenses depuis la feuille DEPENSES du fichier Excel (DATADAF.xlsx)'

    def add_arguments(self, parser):
        parser.add_argument(
            '--file',
            type=str,
            default='',
            help='Chemin vers le fichier Excel (défaut: DATADAF.xlsx à la racine du projet)',
        )
        parser.add_argument(
            '--sheet',
            type=str,
            default='DEPENSES',
            help='Nom de la feuille à lire (défaut: DEPENSES)',
        )
        parser.add_argument(
            '--skip-duplicates',
            action='store_true',
            default=True,
            help='Ne pas insérer les lignes déjà présentes (même date, libellé, banque, montants)',
        )
        parser.add_argument(
            '--dry-run',
            action='store_true',
            help='Afficher ce qui serait importé sans écrire en base',
        )

    def handle(self, *args, **options):
        file_path = options.get('file') or os.path.join(settings.BASE_DIR, 'DATADAF.xlsx')
        sheet_name = options.get('sheet', 'DEPENSES')
        skip_duplicates = options.get('skip_duplicates', True)
        dry_run = options.get('dry_run', False)

        if not os.path.isfile(file_path):
            self.stdout.write(self.style.ERROR(f'Fichier introuvable: {file_path}'))
            return

        try:
            import openpyxl
        except ImportError:
            self.stdout.write(self.style.ERROR('Installer openpyxl: pip install openpyxl'))
            return

        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            self.stdout.write(self.style.ERROR(f'Feuille "{sheet_name}" introuvable. Feuilles: {wb.sheetnames}'))
            wb.close()
            return

        sh = wb[sheet_name]
        # En-têtes ligne 3 : MOIS, ANNEE, DATE, ARTICLE LITTERA, LIBELLE DEPENSES, BANQUE, MONTANT EN Fc, MONTANT EN $us, OBSERVATION
        # Données à partir de la ligne 4
        imported = 0
        skipped = 0
        errors = []

        for row_idx, row in enumerate(sh.iter_rows(min_row=4, values_only=True), start=4):
            if not row or all(c is None for c in row[:9]):
                continue

            try:
                mois_val = row[0]
                annee_val = row[1]
                date_val = row[2]
                article_code = (row[3] or '').strip()[:50] if row[3] is not None else ''
                libelle = (row[4] or '').strip() if row[4] is not None else ''
                banque = (row[5] or '').strip() if row[5] is not None else ''
                montant_fc = self._to_decimal(row[6])
                montant_usd = self._to_decimal(row[7])
                observation = (row[8] or '').strip() if len(row) > 8 and row[8] is not None else ''
            except (IndexError, TypeError) as e:
                errors.append(f'Ligne {row_idx}: {e}')
                continue

            if not libelle and not banque and montant_fc == 0 and montant_usd == 0:
                continue

            if date_val is None:
                if annee_val and mois_val:
                    try:
                        m = int(mois_val) if mois_val is not None else 1
                        a = int(annee_val) if annee_val is not None else 2000
                        date_val = datetime(a, min(max(m, 1), 12), 1).date()
                    except (ValueError, TypeError):
                        errors.append(f'Ligne {row_idx}: date invalide')
                        continue
                else:
                    errors.append(f'Ligne {row_idx}: date manquante')
                    continue

            if hasattr(date_val, 'date'):
                date_val = date_val.date()
            elif isinstance(date_val, str):
                try:
                    date_val = datetime.strptime(date_val[:10], '%Y-%m-%d').date()
                except ValueError:
                    try:
                        date_val = datetime.strptime(date_val[:10], '%d/%m/%Y').date()
                    except ValueError:
                        errors.append(f'Ligne {row_idx}: format date invalide')
                        continue

            try:
                mois = int(mois_val) if mois_val is not None else date_val.month
                annee = int(annee_val) if annee_val is not None else date_val.year
            except (ValueError, TypeError):
                mois = date_val.month
                annee = date_val.year

            mois = max(1, min(12, mois))
            libelle = (libelle or '')[:500]
            banque_nom = (banque or '').strip()[:100]
            observation = (observation or '')[:5000]

            # Résoudre article/code vers NatureEconomique (par code)
            nature_obj = None
            if article_code:
                code_norm = article_code.replace(' ', '').strip()
                nature_obj = NatureEconomique.objects.filter(code=code_norm, active=True).first()
                if not nature_obj:
                    nature_obj = NatureEconomique.objects.filter(code__iexact=article_code.strip(), active=True).first()

            # Résoudre le nom de banque vers la FK Banque
            banque_obj = None
            if banque_nom:
                banque_obj = Banque.objects.filter(nom_banque__iexact=banque_nom).first()
                if not banque_obj:
                    banque_obj = Banque.objects.filter(nom_banque__icontains=banque_nom).first()

            if skip_duplicates and not dry_run:
                exists = DepenseFeuille.objects.filter(
                    date=date_val,
                    libelle_depenses=libelle,
                    banque=banque_obj,
                    montant_fc=montant_fc,
                    montant_usd=montant_usd,
                ).exists()
                if exists:
                    skipped += 1
                    continue

            if dry_run:
                self.stdout.write(
                    f'  [{row_idx}] {date_val} | {libelle[:35]}... | {banque_nom} | FC={montant_fc} | USD={montant_usd}'
                )
                imported += 1
                continue

            try:
                DepenseFeuille.objects.create(
                    mois=mois,
                    annee=annee,
                    date=date_val,
                    nature_economique=nature_obj,
                    libelle_depenses=libelle,
                    banque=banque_obj,
                    montant_fc=montant_fc,
                    montant_usd=montant_usd,
                    observation=observation,
                )
                imported += 1
            except Exception as e:
                errors.append(f'Ligne {row_idx}: {e}')

        wb.close()

        self.stdout.write(self.style.SUCCESS(f'Import terminé: {imported} ligne(s) importée(s), {skipped} doublon(s) ignoré(s).'))
        if errors:
            for err in errors[:20]:
                self.stdout.write(self.style.WARNING(err))
            if len(errors) > 20:
                self.stdout.write(self.style.WARNING(f'... et {len(errors) - 20} autre(s) erreur(s).'))

    def _to_decimal(self, value):
        if value is None or value == '':
            return Decimal('0.00')
        try:
            return Decimal(str(value).replace(',', '.').strip())
        except Exception:
            return Decimal('0.00')
