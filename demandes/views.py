"""
Vues pour la gestion des demandes de paiement
"""
from django.views.generic import ListView, CreateView, UpdateView, DetailView, FormView, View, RedirectView
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.shortcuts import redirect, get_object_or_404, render
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q, Sum
from django.db import transaction
from django.http import JsonResponse, HttpResponse
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from .models import DemandePaiement, ReleveDepense, Depense, NomenclatureDepense, NatureEconomique, Cheque, Paiement
from accounts.models import Service
from banques.models import Banque, CompteBancaire
from releves.models import ReleveBancaire
from .forms import DemandePaiementForm, DemandePaiementValidationForm, ReleveDepenseForm, ReleveDepenseCreateForm, ReleveDepenseAutoForm, DepenseForm, NatureEconomiqueForm, ChequeBanqueForm, PaiementForm, PaiementMultipleForm


class DemandePaiementListView(LoginRequiredMixin, ListView):
    model = DemandePaiement
    template_name = 'demandes/demande_liste.html'
    context_object_name = 'demandes'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = DemandePaiement.objects.select_related(
            'service_demandeur', 'cree_par', 'approuve_par', 'nature_economique'
        ).prefetch_related('releves_depense')
        
        # Filtrage selon le rôle
        if not self.request.user.peut_consulter_tout():
            if self.request.user.is_chef_service:
                queryset = queryset.filter(service_demandeur=self.request.user.service)
        
        # Filtrage par statut
        statut = self.request.GET.get('statut')
        if statut:
            queryset = queryset.filter(statut=statut)
        
        # Par défaut, exclure les demandes déjà dans un relevé
        # Sauf si l'utilisateur demande à voir l'historique complet
        voir_historique = self.request.GET.get('historique', 'false').lower() == 'true'
        if not voir_historique:
            queryset = queryset.exclude(releves_depense__isnull=False)
        
        return queryset.order_by('-date_soumission')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculer les totaux par devise
        queryset = self.get_queryset()
        total_usd = queryset.filter(devise='USD').aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
        total_cdf = queryset.filter(devise='CDF').aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
        
        context['total_usd'] = total_usd
        context['total_cdf'] = total_cdf
        
        return context


class DemandePaiementCreateView(LoginRequiredMixin, CreateView):
    model = DemandePaiement
    form_class = DemandePaiementForm
    template_name = 'demandes/demande_form.html'
    success_url = reverse_lazy('demandes:liste')
    
    def dispatch(self, request, *args, **kwargs):
        # Vérifier que l'utilisateur peut créer une demande (chef de service ou autres rôles autorisés)
        if not request.user.is_chef_service and not request.user.peut_consulter_tout():
            messages.error(request, 'Vous n\'avez pas la permission de créer une demande de paiement. Seuls les chefs de service peuvent créer des demandes.')
            return redirect('demandes:liste')
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        form.instance.cree_par = self.request.user
        messages.success(self.request, 'Demande de paiement créée avec succès.')
        return super().form_valid(form)


class DemandePaiementUpdateView(LoginRequiredMixin, UpdateView):
    model = DemandePaiement
    form_class = DemandePaiementForm
    template_name = 'demandes/demande_form.html'
    success_url = reverse_lazy('demandes:liste')
    
    def dispatch(self, request, *args, **kwargs):
        demande = self.get_object()
        # Vérifier que l'utilisateur est le créateur et que la demande est en attente
        if demande.cree_par != request.user:
            messages.error(request, 'Vous n\'avez pas la permission de modifier cette demande.')
            return redirect('demandes:detail', pk=demande.pk)
        if demande.statut != 'EN_ATTENTE':
            messages.error(request, 'Seules les demandes en attente peuvent être modifiées.')
            return redirect('demandes:detail', pk=demande.pk)
        return super().dispatch(request, *args, **kwargs)
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def form_valid(self, form):
        messages.success(self.request, 'Demande de paiement modifiée avec succès.')
        return super().form_valid(form)


class DemandePaiementDetailView(LoginRequiredMixin, DetailView):
    model = DemandePaiement
    template_name = 'demandes/demande_detail.html'
    context_object_name = 'demande'
    
    def get_queryset(self):
        """Optimiser la requête en préchargeant les relevés associés"""
        return DemandePaiement.objects.select_related(
            'service_demandeur', 'cree_par', 'approuve_par', 'nature_economique', 'nomenclature'
        ).prefetch_related('releves_depense')


class DemandePaiementValidationView(LoginRequiredMixin, FormView):
    model = DemandePaiement
    form_class = DemandePaiementValidationForm
    template_name = 'demandes/demande_validation.html'
    
    def dispatch(self, request, *args, **kwargs):
        self.demande = get_object_or_404(DemandePaiement, pk=kwargs['pk'])
        
        # Vérifier les permissions
        if not request.user.peut_valider_depense():
            messages.error(request, 'Vous n\'avez pas la permission de valider cette demande.')
            return redirect('demandes:detail', pk=self.demande.pk)
        
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['demande'] = self.demande
        return context
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['instance'] = self.demande
        return kwargs
    
    def form_valid(self, form):
        demande = self.demande
        statut = form.cleaned_data['statut']
        
        if statut in ['VALIDEE_DG', 'VALIDEE_DF', 'PAYEE']:
            demande.approuve_par = self.request.user
            demande.date_approbation = timezone.now()
            demande.commentaire_rejet = ''
            messages.success(self.request, f'Demande validée avec succès.')
        elif statut == 'REJETEE':
            demande.commentaire_rejet = form.cleaned_data.get('commentaire_rejet', '')
            messages.warning(self.request, 'Demande rejetée.')
        
        demande.statut = statut
        demande.save()
        
        return redirect('demandes:detail', pk=demande.pk)


class ReleveDepenseListView(LoginRequiredMixin, RedirectView):
    """Vue pour rediriger vers la liste des relevés créés"""
    permanent = False
    pattern_name = 'demandes:releves_crees_liste'
    query_string = True


class ReleveDepenseGenererPDFView(LoginRequiredMixin, DetailView):
    """Vue pour générer un PDF d'un relevé de dépense existant"""
    model = ReleveDepense
    template_name = 'demandes/releve_pdf_template.html'
    context_object_name = 'releve'
    
    def dispatch(self, request, *args, **kwargs):
        # Vérifier les permissions
        if not request.user.peut_consulter_tout():
            messages.error(request, 'Vous n\'avez pas la permission de générer un PDF de relevé.')
            return redirect('demandes:releves_liste')
        return super().dispatch(request, *args, **kwargs)
    
    def get(self, request, *args, **kwargs):
        releve = self.get_object()
        
        # Créer le PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="releve_depense_{releve.numero}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        # Créer le document PDF
        doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
        
        # Contenu du PDF
        story = []
        styles = getSampleStyleSheet()
        
        # Style personnalisé pour les titres
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.darkblue
        )
        
        # Style pour les en-têtes de tableau
        header_style = ParagraphStyle(
            'TableHeader',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            textColor=colors.white
        )
        
        # En-tête
        story.append(Paragraph("DGRAD - DIRECTION GÉNÉRALE DES REVENUS", title_style))
        story.append(Paragraph("RELEVÉ DES DÉPENSES", styles['Heading2']))
        story.append(Paragraph(f"Numéro: {releve.numero}", styles['Normal']))
        story.append(Paragraph(f"Période: {releve.periode}", styles['Normal']))
        story.append(Paragraph(f"Date de création: {releve.date_creation.strftime('%d/%m/%Y')}", styles['Normal']))
        story.append(Spacer(1, 20))
        
        # Tableau des demandes
        demandes = releve.demandes.filter(
            statut__in=['VALIDEE_DG', 'VALIDEE_DF', 'PAYEE']
        ).select_related('service_demandeur', 'nature_economique').order_by('nature_economique__code')
        
        if demandes:
            # En-têtes du tableau
            headers = ['Référence', 'Service', 'Nature Éco', 'Description', 'Montant', 'Devise']
            data = [headers]
            
            # Données
            for demande in demandes:
                data.append([
                    demande.reference,
                    demande.service_demandeur.nom_service if demande.service_demandeur else '',
                    demande.nature_economique.code if demande.nature_economique else '',
                    demande.description[:50] + '...' if len(demande.description) > 50 else demande.description,
                    str(demande.montant),
                    demande.devise
                ])
            
            # Créer le tableau
            table = Table(data)
            table.setStyle(TableStyle([
                # En-tête
                ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                
                # Données
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('ALIGN', (4, 1), (5, -1), 'RIGHT'),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                
                # Bordures
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            ]))
            
            story.append(table)
            story.append(Spacer(1, 20))
        
        # Totaux
        story.append(Paragraph("RÉCAPITULATIF", styles['Heading3']))
        
        totaux_data = [
            ['Description', 'Montant CDF', 'Montant USD'],
            ['Total Brut', f"{releve.montant_cdf:,.2f}", f"{releve.montant_usd:,.2f}"],
            ['IPR (3%)', f"{releve.ipr_cdf:,.2f}", f"{releve.ipr_usd:,.2f}"],
            ['Net à Payer', f"{releve.net_a_payer_cdf:,.2f}", f"{releve.net_a_payer_usd:,.2f}"],
        ]
        
        totaux_table = Table(totaux_data)
        totaux_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        
        story.append(totaux_table)
        
        # Observation
        if releve.observation:
            story.append(Spacer(1, 20))
            story.append(Paragraph("OBSERVATION", styles['Heading3']))
            story.append(Paragraph(releve.observation, styles['Normal']))
        
        # Signatures
        story.append(Spacer(1, 30))
        story.append(Paragraph("Validé par:", styles['Normal']))
        story.append(Paragraph(f"{releve.valide_par.get_full_name() or releve.valide_par.username}", styles['Normal']))
        
        # Construire le PDF
        doc.build(story)
        
        return response


class ReleveDepensePDFView(LoginRequiredMixin, ListView):
    """Vue pour générer un PDF du relevé de dépense"""
    model = DemandePaiement
    
    def get_queryset(self):
        """Récupérer uniquement les demandes validées qui ne sont pas déjà dans un relevé"""
        queryset = DemandePaiement.objects.select_related(
            'service_demandeur', 'cree_par', 'approuve_par', 'nature_economique'
        ).filter(
            statut__in=['VALIDEE_DG', 'VALIDEE_DF', 'PAYEE']
        ).exclude(
            releves_depense__isnull=False  # Exclure les demandes déjà dans un relevé
        )
        
        # Filtrage selon le rôle
        if not self.request.user.peut_consulter_tout():
            if self.request.user.is_chef_service:
                queryset = queryset.filter(service_demandeur=self.request.user.service)
        
        # Trier par code de nature économique, puis par date
        return queryset.order_by('nature_economique__code', '-date_soumission')
    
    def get_context_data(self, **kwargs):
        context = {}
        
        # Calculer les totaux pour toutes les demandes validées
        queryset = self.get_queryset()
        
        # Montants par devise
        montant_cdf = queryset.filter(devise='CDF').aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
        montant_usd = queryset.filter(devise='USD').aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
        
        # IPR (3%)
        ipr_cdf = montant_cdf * Decimal('0.03')
        ipr_usd = montant_usd * Decimal('0.03')
        
        # Net à payer (montant - IPR)
        net_a_payer_cdf = montant_cdf - ipr_cdf
        net_a_payer_usd = montant_usd - ipr_usd
        
        context['montant_cdf'] = montant_cdf
        context['montant_usd'] = montant_usd
        context['ipr_cdf'] = ipr_cdf
        context['ipr_usd'] = ipr_usd
        context['net_a_payer_cdf'] = net_a_payer_cdf
        context['net_a_payer_usd'] = net_a_payer_usd
        
        # Grouper les demandes par code de nature économique pour les sous-totaux
        from collections import defaultdict
        demandes_par_code = defaultdict(list)
        demande_numero = {}
        numero = 1
        
        for demande in queryset:
            code = demande.nature_economique.code if demande.nature_economique else 'Sans code'
            demandes_par_code[code].append(demande)
            demande_numero[demande.pk] = numero
            numero += 1
        
        # Calculer les sous-totaux pour chaque groupe
        sous_totaux = {}
        for code, demandes_groupe in demandes_par_code.items():
            montant_usd_groupe = sum(d.montant for d in demandes_groupe if d.devise == 'USD')
            montant_cdf_groupe = sum(d.montant for d in demandes_groupe if d.devise == 'CDF')
            ipr_usd_groupe = montant_usd_groupe * Decimal('0.03')
            ipr_cdf_groupe = montant_cdf_groupe * Decimal('0.03')
            net_usd_groupe = montant_usd_groupe - ipr_usd_groupe
            net_cdf_groupe = montant_cdf_groupe - ipr_cdf_groupe
            
            sous_totaux[code] = {
                'montant_usd': montant_usd_groupe,
                'montant_cdf': montant_cdf_groupe,
                'ipr_usd': ipr_usd_groupe,
                'ipr_cdf': ipr_cdf_groupe,
                'net_usd': net_usd_groupe,
                'net_cdf': net_cdf_groupe,
            }
        
        context['sous_totaux'] = sous_totaux
        context['demande_numero'] = demande_numero
        context['demandes_par_code'] = demandes_par_code
        context['queryset'] = queryset
        
        return context
    
    def get(self, request, *args, **kwargs):
        context = self.get_context_data()
        queryset = context['queryset']
        demandes_par_code = context['demandes_par_code']
        
        # Créer ou récupérer un relevé pour cette période (mois en cours)
        periode = timezone.now().date().replace(day=1)  # Premier jour du mois
        
        # Utiliser une transaction pour éviter les conflits en cas de requêtes simultanées
        with transaction.atomic():
            releve, created = ReleveDepense.objects.get_or_create(
                periode=periode,
                defaults={
                    'valide_par': request.user,
                    'observation': f'Relevé généré automatiquement le {timezone.now().strftime("%d/%m/%Y")}'
                }
            )
            
            # S'assurer que le relevé a un numéro (au cas où il aurait été créé sans)
            if not releve.numero:
                releve.save()  # Cela générera le numéro
            
            # Si le relevé vient d'être créé, ajouter les demandes et calculer les totaux
            if created:
                # Utiliser la méthode sécurisée pour éviter les doublons
                try:
                    releve.ajouter_demandes_securise(queryset)
                    releve.calculer_total()
                except Exception as e:
                    messages.error(request, f'Erreur lors de l\'ajout des demandes : {str(e)}')
                    return redirect('demandes:releves_liste')
            else:
                # Si le relevé existe déjà, ne pas modifier les demandes
                # Juste recalculer les totaux avec uniquement les demandes validées
                releve.calculer_total()
        
        # Créer le PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="releve_depense_{releve.numero}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        # Créer le document PDF en mode paysage
        doc = SimpleDocTemplate(response, pagesize=landscape(A4), 
                                rightMargin=1*cm, leftMargin=1*cm, 
                                topMargin=1*cm, bottomMargin=1*cm)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a3a5f'),
            alignment=TA_CENTER,
            spaceAfter=30,
        )
        
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_RIGHT,
            spaceBefore=20,
        )
        
        # Titre avec numéro de relevé
        date_du_jour = releve.date_creation.strftime('%d/%m/%Y')
        title = Paragraph(f"RELEVÉ DE DÉPENSE N° {releve.numero} DU {date_du_jour}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*cm))
        
              
        # Résumé des montants
        resume_data = [
            ['Désignation', 'CDF', 'USD'],
            ['Montant', f"{context['montant_cdf']:,.2f}".replace(',', ' ').replace('.', ','), 
                      f"{context['montant_usd']:,.2f}".replace(',', ' ').replace('.', ',')],
            ['IPR (3%)', f"{context['ipr_cdf']:,.2f}".replace(',', ' ').replace('.', ','), 
                      f"{context['ipr_usd']:,.2f}".replace(',', ' ').replace('.', ',')],
            ['Net à payer', f"{context['net_a_payer_cdf']:,.2f}".replace(',', ' ').replace('.', ','), 
                      f"{context['net_a_payer_usd']:,.2f}".replace(',', ' ').replace('.', ',')],
        ]
        
        resume_table = Table(resume_data, colWidths=[6*cm, 4*cm, 4*cm])
        resume_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        elements.append(resume_table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Tableau principal
        # En-têtes
        table_data = [['N°', 'Code', 'Nature économique', 'Montant USD', 'IPR USD', 
                      'Montant CDF', 'IPR CDF', 'Net à payer USD', 'Net à payer CDF']]
        
        numero = 1
        for code in sorted(demandes_par_code.keys()):
            demandes_groupe = demandes_par_code[code]
            for demande in demandes_groupe:
                montant_usd = demande.montant if demande.devise == 'USD' else Decimal('0.00')
                montant_cdf = demande.montant if demande.devise == 'CDF' else Decimal('0.00')
                ipr_usd = montant_usd * Decimal('0.03')
                ipr_cdf = montant_cdf * Decimal('0.03')
                net_usd = montant_usd - ipr_usd
                net_cdf = montant_cdf - ipr_cdf
                
                nature = demande.nature_economique.titre if demande.nature_economique else '-'
                if len(nature) > 30:
                    nature = nature[:27] + '...'
                
                table_data.append([
                    str(numero),
                    code,
                    nature,
                    f"{montant_usd:,.2f}".replace(',', ' ').replace('.', ',') if montant_usd > 0 else '-',
                    f"{ipr_usd:,.2f}".replace(',', ' ').replace('.', ',') if ipr_usd > 0 else '-',
                    f"{montant_cdf:,.2f}".replace(',', ' ').replace('.', ',') if montant_cdf > 0 else '-',
                    f"{ipr_cdf:,.2f}".replace(',', ' ').replace('.', ',') if ipr_cdf > 0 else '-',
                    f"{net_usd:,.2f}".replace(',', ' ').replace('.', ',') if net_usd > 0 else '-',
                    f"{net_cdf:,.2f}".replace(',', ' ').replace('.', ',') if net_cdf > 0 else '-',
                ])
                numero += 1
            
            # Sous-total
            if code in context['sous_totaux']:
                total = context['sous_totaux'][code]
                table_data.append([
                    '', '', f'Sous-total (Code: {code})',
                    f"{total['montant_usd']:,.2f}".replace(',', ' ').replace('.', ',') if total['montant_usd'] > 0 else '-',
                    f"{total['ipr_usd']:,.2f}".replace(',', ' ').replace('.', ',') if total['ipr_usd'] > 0 else '-',
                    f"{total['montant_cdf']:,.2f}".replace(',', ' ').replace('.', ',') if total['montant_cdf'] > 0 else '-',
                    f"{total['ipr_cdf']:,.2f}".replace(',', ' ').replace('.', ',') if total['ipr_cdf'] > 0 else '-',
                    f"{total['net_usd']:,.2f}".replace(',', ' ').replace('.', ',') if total['net_usd'] > 0 else '-',
                    f"{total['net_cdf']:,.2f}".replace(',', ' ').replace('.', ',') if total['net_cdf'] > 0 else '-',
                ])
        
        # Total général
        table_data.append([
            '', '', 'TOTAL GÉNÉRAL',
            f"{context['montant_usd']:,.2f}".replace(',', ' ').replace('.', ',') if context['montant_usd'] > 0 else '-',
            f"{context['ipr_usd']:,.2f}".replace(',', ' ').replace('.', ',') if context['ipr_usd'] > 0 else '-',
            f"{context['montant_cdf']:,.2f}".replace(',', ' ').replace('.', ',') if context['montant_cdf'] > 0 else '-',
            f"{context['ipr_cdf']:,.2f}".replace(',', ' ').replace('.', ',') if context['ipr_cdf'] > 0 else '-',
            f"{context['net_a_payer_usd']:,.2f}".replace(',', ' ').replace('.', ',') if context['net_a_payer_usd'] > 0 else '-',
            f"{context['net_a_payer_cdf']:,.2f}".replace(',', ' ').replace('.', ',') if context['net_a_payer_cdf'] > 0 else '-',
        ])
        
        # Créer le tableau
        table = Table(table_data, colWidths=[1*cm, 2*cm, 4*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -2), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
            # Style pour les sous-totaux
            ('BACKGROUND', (0, -2), (-1, -2), colors.lightgrey),
            ('FONTNAME', (0, -2), (-1, -2), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -2), (-1, -2), 8),
            # Style pour le total général
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1a3a5f')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 9),
        ]))
        
        
        elements.append(table)
        
        # Ajouter "Fait à Kinshasa" juste après le tableau, aligné à droite avec la même largeur que le tableau
        # Largeur totale du tableau : 1 + 2 + 4 + 2.5*6 = 22 cm
        table_width = 1*cm + 2*cm + 4*cm + 2.5*cm*6
        date_text = f"Fait à Kinshasa le {timezone.now().strftime('%d/%m/%Y')}"
        footer_paragraph = Paragraph(date_text, footer_style)
        # Créer un tableau avec une seule cellule de la même largeur que le tableau principal
        footer_table = Table([[footer_paragraph]], colWidths=[table_width])
        footer_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(Spacer(1, 0.5*cm))
        elements.append(footer_table)
        
        # Ajouter les trois signataires
        elements.append(Spacer(1, 1*cm))
        
        # Styles pour les signataires
        signataire_gauche_style = ParagraphStyle(
            'SignataireGaucheStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_LEFT,
        )
        
        signataire_droite_style = ParagraphStyle(
            'SignataireDroiteStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_RIGHT,
        )
        
        signataire_centre_style = ParagraphStyle(
            'SignataireCentreStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
        )
        
        # Première ligne : deux signataires (gauche et droite)
        signataire_gauche = Paragraph("_________________________<br/>Signature 1", signataire_gauche_style)
        signataire_droite = Paragraph("_________________________<br/>Signature 2", signataire_droite_style)
        
        # Créer un tableau pour la première ligne avec deux signataires
        signataires_ligne1_data = [[signataire_gauche, signataire_droite]]
        signataires_ligne1_table = Table(signataires_ligne1_data, colWidths=[table_width/2, table_width/2])
        signataires_ligne1_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),  # Premier signataire à gauche
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),  # Deuxième signataire à droite
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        
        elements.append(signataires_ligne1_table)
        
        # Deuxième ligne : un signataire centré un peu plus bas
        elements.append(Spacer(1, 0.8*cm))
        signataire_centre = Paragraph("_________________________<br/>Signature 3", signataire_centre_style)
        
        # Créer un tableau pour la deuxième ligne avec un signataire centré
        signataires_ligne2_data = [[signataire_centre]]
        signataires_ligne2_table = Table(signataires_ligne2_data, colWidths=[table_width])
        signataires_ligne2_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Centrer le troisième signataire
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        
        elements.append(signataires_ligne2_table)
        
        # Construire le PDF
        doc.build(elements)
        
        return response


class ReleveDepenseExcelView(LoginRequiredMixin, ListView):
    """Vue pour générer un fichier Excel du relevé de dépense avec la même mise en forme que le PDF"""
    model = DemandePaiement
    
    def get_queryset(self):
        """Récupérer uniquement les demandes validées qui ne sont pas déjà dans un relevé"""
        queryset = DemandePaiement.objects.select_related(
            'service_demandeur', 'cree_par', 'approuve_par', 'nature_economique'
        ).filter(
            statut__in=['VALIDEE_DG', 'VALIDEE_DF', 'PAYEE']
        ).exclude(
            releves_depense__isnull=False  # Exclure les demandes déjà dans un relevé
        )
        
        # Filtrage selon le rôle
        if not self.request.user.peut_consulter_tout():
            if self.request.user.is_chef_service:
                queryset = queryset.filter(service_demandeur=self.request.user.service)
        
        # Trier par code de nature économique, puis par date
        return queryset.order_by('nature_economique__code', '-date_soumission')
    
    def get_context_data(self, **kwargs):
        context = {}
        
        # Calculer les totaux pour toutes les demandes validées
        queryset = self.get_queryset()
        
        # Montants par devise
        montant_cdf = queryset.filter(devise='CDF').aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
        montant_usd = queryset.filter(devise='USD').aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
        
        # IPR (3%)
        ipr_cdf = montant_cdf * Decimal('0.03')
        ipr_usd = montant_usd * Decimal('0.03')
        
        # Net à payer (montant - IPR)
        net_a_payer_cdf = montant_cdf - ipr_cdf
        net_a_payer_usd = montant_usd - ipr_usd
        
        context['montant_cdf'] = montant_cdf
        context['montant_usd'] = montant_usd
        context['ipr_cdf'] = ipr_cdf
        context['ipr_usd'] = ipr_usd
        context['net_a_payer_cdf'] = net_a_payer_cdf
        context['net_a_payer_usd'] = net_a_payer_usd
        
        # Grouper les demandes par code de nature économique pour les sous-totaux
        from collections import defaultdict
        demandes_par_code = defaultdict(list)
        demande_numero = {}
        numero = 1
        
        for demande in queryset:
            code = demande.nature_economique.code if demande.nature_economique else 'Sans code'
            demandes_par_code[code].append(demande)
            demande_numero[demande.pk] = numero
            numero += 1
        
        # Calculer les sous-totaux pour chaque groupe
        sous_totaux = {}
        for code, demandes_groupe in demandes_par_code.items():
            montant_usd_groupe = sum(d.montant for d in demandes_groupe if d.devise == 'USD')
            montant_cdf_groupe = sum(d.montant for d in demandes_groupe if d.devise == 'CDF')
            ipr_usd_groupe = montant_usd_groupe * Decimal('0.03')
            ipr_cdf_groupe = montant_cdf_groupe * Decimal('0.03')
            net_usd_groupe = montant_usd_groupe - ipr_usd_groupe
            net_cdf_groupe = montant_cdf_groupe - ipr_cdf_groupe
            
            sous_totaux[code] = {
                'montant_usd': montant_usd_groupe,
                'montant_cdf': montant_cdf_groupe,
                'ipr_usd': ipr_usd_groupe,
                'ipr_cdf': ipr_cdf_groupe,
                'net_usd': net_usd_groupe,
                'net_cdf': net_cdf_groupe,
            }
        
        context['sous_totaux'] = sous_totaux
        context['demande_numero'] = demande_numero
        context['demandes_par_code'] = demandes_par_code
        context['queryset'] = queryset
        
        return context
    
    def get(self, request, *args, **kwargs):
        context = self.get_context_data()
        queryset = context['queryset']
        demandes_par_code = context['demandes_par_code']
        
        # Créer le fichier Excel
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="releve_depense_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        
        # Créer le workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Relevé de Dépense"
        
        # Styles
        header_fill = PatternFill(start_color="1a3a5f", end_color="1a3a5f", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        title_font = Font(bold=True, size=16, color="1a3a5f")
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        
        row = 1
        
        # Titre
        date_du_jour = timezone.now().strftime('%d/%m/%Y')
        ws.merge_cells(f'A{row}:I{row}')
        cell = ws.cell(row=row, column=1, value=f"RELEVÉ DE DÉPENSE DU {date_du_jour}")
        cell.font = title_font
        cell.alignment = center_align
        row += 2
        
        # Résumé des montants
        resume_start_row = row
        resume_headers = ['Désignation', 'CDF', 'USD']
        for col, header in enumerate(resume_headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        
        row += 1
        resume_data = [
            ['Montant', context['montant_cdf'], context['montant_usd']],
            ['IPR (3%)', context['ipr_cdf'], context['ipr_usd']],
            ['Net à payer', context['net_a_payer_cdf'], context['net_a_payer_usd']],
        ]
        
        for data_row in resume_data:
            for col, value in enumerate(data_row, start=1):
                cell = ws.cell(row=row, column=col, value=value if col == 1 else float(value))
                cell.alignment = center_align if col == 1 else right_align
                cell.border = border
                if col > 1:
                    cell.number_format = '#,##0.00'
            row += 1
        
        # Appliquer le style beige aux données du résumé
        for r in range(resume_start_row + 1, row):
            for c in range(1, 4):
                ws.cell(row=r, column=c).fill = PatternFill(start_color="F5F5DC", end_color="F5F5DC", fill_type="solid")
        
        row += 2
        
        # Tableau principal - En-têtes
        headers = ['N°', 'Code', 'Nature économique', 'Montant USD', 'IPR USD', 
                   'Montant CDF', 'IPR CDF', 'Net à payer USD', 'Net à payer CDF']
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = border
        row += 1
        
        # Données du tableau
        numero = 1
        for code in sorted(demandes_par_code.keys()):
            demandes_groupe = demandes_par_code[code]
            for demande in demandes_groupe:
                montant_usd = demande.montant if demande.devise == 'USD' else Decimal('0.00')
                montant_cdf = demande.montant if demande.devise == 'CDF' else Decimal('0.00')
                ipr_usd = montant_usd * Decimal('0.03')
                ipr_cdf = montant_cdf * Decimal('0.03')
                net_usd = montant_usd - ipr_usd
                net_cdf = montant_cdf - ipr_cdf
                
                nature = demande.nature_economique.titre if demande.nature_economique else '-'
                
                data = [
                    numero,
                    code,
                    nature,
                    float(montant_usd) if montant_usd > 0 else '-',
                    float(ipr_usd) if ipr_usd > 0 else '-',
                    float(montant_cdf) if montant_cdf > 0 else '-',
                    float(ipr_cdf) if ipr_cdf > 0 else '-',
                    float(net_usd) if net_usd > 0 else '-',
                    float(net_cdf) if net_cdf > 0 else '-',
                ]
                
                for col, value in enumerate(data, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = center_align if col <= 3 else right_align
                    cell.border = border
                    if col > 3 and isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                    # Alterner les couleurs de fond
                    if row % 2 == 0:
                        cell.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
                
                row += 1
                numero += 1
            
            # Sous-total
            if code in context['sous_totaux']:
                total = context['sous_totaux'][code]
                data = [
                    '', '', f'Sous-total (Code: {code})',
                    float(total['montant_usd']) if total['montant_usd'] > 0 else '-',
                    float(total['ipr_usd']) if total['ipr_usd'] > 0 else '-',
                    float(total['montant_cdf']) if total['montant_cdf'] > 0 else '-',
                    float(total['ipr_cdf']) if total['ipr_cdf'] > 0 else '-',
                    float(total['net_usd']) if total['net_usd'] > 0 else '-',
                    float(total['net_cdf']) if total['net_cdf'] > 0 else '-',
                ]
                
                for col, value in enumerate(data, start=1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.alignment = center_align if col <= 3 else right_align
                    cell.border = border
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                    if col > 3 and isinstance(value, (int, float)):
                        cell.number_format = '#,##0.00'
                row += 1
        
        # Total général
        data = [
            '', '', 'TOTAL GÉNÉRAL',
            float(context['montant_usd']) if context['montant_usd'] > 0 else '-',
            float(context['ipr_usd']) if context['ipr_usd'] > 0 else '-',
            float(context['montant_cdf']) if context['montant_cdf'] > 0 else '-',
            float(context['ipr_cdf']) if context['ipr_cdf'] > 0 else '-',
            float(context['net_a_payer_usd']) if context['net_a_payer_usd'] > 0 else '-',
            float(context['net_a_payer_cdf']) if context['net_a_payer_cdf'] > 0 else '-',
        ]
        
        for col, value in enumerate(data, start=1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.alignment = center_align if col <= 3 else right_align
            cell.border = border
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = header_fill
            if col > 3 and isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
        
        row += 2
        
        # "Fait à Kinshasa" aligné à droite
        ws.merge_cells(f'A{row}:I{row}')
        date_text = f"Fait à Kinshasa le {timezone.now().strftime('%d/%m/%Y')}"
        cell = ws.cell(row=row, column=1, value=date_text)
        cell.alignment = Alignment(horizontal='right', vertical='center')
        cell.font = Font(size=10)
        row += 3
        
        # Signataires
        # Première ligne : deux signataires (gauche et droite)
        ws.merge_cells(f'A{row}:E{row}')
        ws.merge_cells(f'F{row}:I{row}')
        cell_left = ws.cell(row=row, column=1, value="_________________________\nSignature 1")
        cell_left.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        cell_right = ws.cell(row=row, column=6, value="_________________________\nSignature 2")
        cell_right.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
        row += 3
        
        # Deuxième ligne : signataire centré
        ws.merge_cells(f'A{row}:I{row}')
        cell_center = ws.cell(row=row, column=1, value="_________________________\nSignature 3")
        cell_center.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        # Ajuster les largeurs des colonnes
        ws.column_dimensions['A'].width = 8
        ws.column_dimensions['B'].width = 12
        ws.column_dimensions['C'].width = 30
        for col in range(4, 10):
            ws.column_dimensions[get_column_letter(col)].width = 15
        
        # Sauvegarder le workbook
        wb.save(response)
        
        return response


class ReleveDepenseReprintPDFView(LoginRequiredMixin, View):
    """Vue pour réimprimer un PDF du relevé de dépense à partir de son numéro"""
    
    def get(self, request, *args, **kwargs):
        numero = request.GET.get('numero')
        if not numero:
            messages.error(request, 'Veuillez fournir un numéro de relevé.')
            return redirect('demandes:releves_liste')
        
        try:
            releve = ReleveDepense.objects.get(numero=numero)
        except ReleveDepense.DoesNotExist:
            messages.error(request, f'Relevé avec le numéro {numero} introuvable.')
            return redirect('demandes:releves_liste')
        
        # Récupérer les demandes du relevé
        demandes = releve.demandes.all().select_related(
            'service_demandeur', 'cree_par', 'approuve_par', 'nature_economique'
        ).order_by('nature_economique__code', '-date_soumission')
        
        # Calculer les totaux
        montant_cdf = sum(d.montant for d in demandes if d.devise == 'CDF')
        montant_usd = sum(d.montant for d in demandes if d.devise == 'USD')
        ipr_cdf = montant_cdf * Decimal('0.03')
        ipr_usd = montant_usd * Decimal('0.03')
        net_a_payer_cdf = montant_cdf - ipr_cdf
        net_a_payer_usd = montant_usd - ipr_usd
        
        # Grouper les demandes par code
        from collections import defaultdict
        demandes_par_code = defaultdict(list)
        demande_numero = {}
        numero_demande = 1
        
        for demande in demandes:
            code = demande.nature_economique.code if demande.nature_economique else 'Sans code'
            demandes_par_code[code].append(demande)
            demande_numero[demande.pk] = numero_demande
            numero_demande += 1
        
        # Calculer les sous-totaux
        sous_totaux = {}
        for code, demandes_groupe in demandes_par_code.items():
            montant_usd_groupe = sum(d.montant for d in demandes_groupe if d.devise == 'USD')
            montant_cdf_groupe = sum(d.montant for d in demandes_groupe if d.devise == 'CDF')
            ipr_usd_groupe = montant_usd_groupe * Decimal('0.03')
            ipr_cdf_groupe = montant_cdf_groupe * Decimal('0.03')
            net_usd_groupe = montant_usd_groupe - ipr_usd_groupe
            net_cdf_groupe = montant_cdf_groupe - ipr_cdf_groupe
            
            sous_totaux[code] = {
                'montant_usd': montant_usd_groupe,
                'montant_cdf': montant_cdf_groupe,
                'ipr_usd': ipr_usd_groupe,
                'ipr_cdf': ipr_cdf_groupe,
                'net_usd': net_usd_groupe,
                'net_cdf': net_cdf_groupe,
            }
        
        # Créer le PDF (utiliser le même code que ReleveDepensePDFView)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="releve_depense_{releve.numero}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=landscape(A4), 
                                rightMargin=1*cm, leftMargin=1*cm, 
                                topMargin=1*cm, bottomMargin=1*cm)
        elements = []
        
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a3a5f'),
            alignment=TA_CENTER,
            spaceAfter=30,
        )
        
        footer_style = ParagraphStyle(
            'FooterStyle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_RIGHT,
            spaceBefore=20,
        )
        
        # Titre avec numéro de relevé
        date_du_jour = releve.date_creation.strftime('%d/%m/%Y')
        title = Paragraph(f"RELEVÉ DE DÉPENSE N° {releve.numero} DU {date_du_jour}", title_style)
        elements.append(title)
        elements.append(Spacer(1, 0.3*cm))
              
        # Résumé des montants
        resume_data = [
            ['Désignation', 'CDF', 'USD'],
            ['Montant', f"{montant_cdf:,.2f}".replace(',', ' ').replace('.', ','), 
                      f"{montant_usd:,.2f}".replace(',', ' ').replace('.', ',')],
            ['IPR (3%)', f"{ipr_cdf:,.2f}".replace(',', ' ').replace('.', ','), 
                      f"{ipr_usd:,.2f}".replace(',', ' ').replace('.', ',')],
            ['Net à payer', f"{net_a_payer_cdf:,.2f}".replace(',', ' ').replace('.', ','), 
                      f"{net_a_payer_usd:,.2f}".replace(',', ' ').replace('.', ',')],
        ]
        
        resume_table = Table(resume_data, colWidths=[6*cm, 4*cm, 4*cm])
        resume_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        elements.append(resume_table)
        elements.append(Spacer(1, 0.5*cm))
        
        # Tableau principal
        table_data = [['N°', 'Code', 'Nature économique', 'Montant USD', 'IPR USD', 
                      'Montant CDF', 'IPR CDF', 'Net à payer USD', 'Net à payer CDF']]
        
        numero_dem = 1
        for code in sorted(demandes_par_code.keys()):
            demandes_groupe = demandes_par_code[code]
            for demande in demandes_groupe:
                montant_usd_d = demande.montant if demande.devise == 'USD' else Decimal('0.00')
                montant_cdf_d = demande.montant if demande.devise == 'CDF' else Decimal('0.00')
                ipr_usd_d = montant_usd_d * Decimal('0.03')
                ipr_cdf_d = montant_cdf_d * Decimal('0.03')
                net_usd_d = montant_usd_d - ipr_usd_d
                net_cdf_d = montant_cdf_d - ipr_cdf_d
                
                nature = demande.nature_economique.titre if demande.nature_economique else '-'
                if len(nature) > 30:
                    nature = nature[:27] + '...'
                
                table_data.append([
                    str(numero_dem),
                    code,
                    nature,
                    f"{montant_usd_d:,.2f}".replace(',', ' ').replace('.', ',') if montant_usd_d > 0 else '-',
                    f"{ipr_usd_d:,.2f}".replace(',', ' ').replace('.', ',') if ipr_usd_d > 0 else '-',
                    f"{montant_cdf_d:,.2f}".replace(',', ' ').replace('.', ',') if montant_cdf_d > 0 else '-',
                    f"{ipr_cdf_d:,.2f}".replace(',', ' ').replace('.', ',') if ipr_cdf_d > 0 else '-',
                    f"{net_usd_d:,.2f}".replace(',', ' ').replace('.', ',') if net_usd_d > 0 else '-',
                    f"{net_cdf_d:,.2f}".replace(',', ' ').replace('.', ',') if net_cdf_d > 0 else '-',
                ])
                numero_dem += 1
            
            # Sous-total
            if code in sous_totaux:
                total = sous_totaux[code]
                table_data.append([
                    '', '', f'Sous-total (Code: {code})',
                    f"{total['montant_usd']:,.2f}".replace(',', ' ').replace('.', ',') if total['montant_usd'] > 0 else '-',
                    f"{total['ipr_usd']:,.2f}".replace(',', ' ').replace('.', ',') if total['ipr_usd'] > 0 else '-',
                    f"{total['montant_cdf']:,.2f}".replace(',', ' ').replace('.', ',') if total['montant_cdf'] > 0 else '-',
                    f"{total['ipr_cdf']:,.2f}".replace(',', ' ').replace('.', ',') if total['ipr_cdf'] > 0 else '-',
                    f"{total['net_usd']:,.2f}".replace(',', ' ').replace('.', ',') if total['net_usd'] > 0 else '-',
                    f"{total['net_cdf']:,.2f}".replace(',', ' ').replace('.', ',') if total['net_cdf'] > 0 else '-',
                ])
        
        # Total général
        table_data.append([
            '', '', 'TOTAL GÉNÉRAL',
            f"{montant_usd:,.2f}".replace(',', ' ').replace('.', ',') if montant_usd > 0 else '-',
            f"{ipr_usd:,.2f}".replace(',', ' ').replace('.', ',') if ipr_usd > 0 else '-',
            f"{montant_cdf:,.2f}".replace(',', ' ').replace('.', ',') if montant_cdf > 0 else '-',
            f"{ipr_cdf:,.2f}".replace(',', ' ').replace('.', ',') if ipr_cdf > 0 else '-',
            f"{net_a_payer_usd:,.2f}".replace(',', ' ').replace('.', ',') if net_a_payer_usd > 0 else '-',
            f"{net_a_payer_cdf:,.2f}".replace(',', ' ').replace('.', ',') if net_a_payer_cdf > 0 else '-',
        ])
        
        table = Table(table_data, colWidths=[1*cm, 2*cm, 4*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm, 2.5*cm])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('ALIGN', (3, 1), (-1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -2), 7),
            ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.lightgrey]),
            ('BACKGROUND', (0, -2), (-1, -2), colors.lightgrey),
            ('FONTNAME', (0, -2), (-1, -2), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -2), (-1, -2), 8),
            ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#1a3a5f')),
            ('TEXTCOLOR', (0, -1), (-1, -1), colors.whitesmoke),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -1), (-1, -1), 9),
        ]))
        
        elements.append(table)
        
        # "Fait à Kinshasa" aligné à droite
        table_width = 1*cm + 2*cm + 4*cm + 2.5*cm*6
        date_text = f"Fait à Kinshasa le {timezone.now().strftime('%d/%m/%Y')}"
        footer_paragraph = Paragraph(date_text, footer_style)
        footer_table = Table([[footer_paragraph]], colWidths=[table_width])
        footer_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(Spacer(1, 0.5*cm))
        elements.append(footer_table)
        
        # Signataires
        elements.append(Spacer(1, 1*cm))
        signataire_gauche_style = ParagraphStyle('SignataireGaucheStyle', parent=styles['Normal'], fontSize=10, alignment=TA_LEFT)
        signataire_droite_style = ParagraphStyle('SignataireDroiteStyle', parent=styles['Normal'], fontSize=10, alignment=TA_RIGHT)
        signataire_centre_style = ParagraphStyle('SignataireCentreStyle', parent=styles['Normal'], fontSize=10, alignment=TA_CENTER)
        
        signataire_gauche = Paragraph("_________________________<br/>Signature 1", signataire_gauche_style)
        signataire_droite = Paragraph("_________________________<br/>Signature 2", signataire_droite_style)
        signataires_ligne1_data = [[signataire_gauche, signataire_droite]]
        signataires_ligne1_table = Table(signataires_ligne1_data, colWidths=[table_width/2, table_width/2])
        signataires_ligne1_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, 0), 'LEFT'),
            ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(signataires_ligne1_table)
        elements.append(Spacer(1, 0.8*cm))
        signataire_centre = Paragraph("_________________________<br/>Signature 3", signataire_centre_style)
        signataires_ligne2_data = [[signataire_centre]]
        signataires_ligne2_table = Table(signataires_ligne2_data, colWidths=[table_width])
        signataires_ligne2_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 0),
        ]))
        elements.append(signataires_ligne2_table)
        
        doc.build(elements)
        return response


class ReleveDepenseListCreatedView(LoginRequiredMixin, ListView):
    """Vue pour afficher la liste des relevés de dépense créés"""
    model = ReleveDepense
    template_name = 'demandes/releve_created_list.html'
    context_object_name = 'releves'
    paginate_by = 20
    
    def get_queryset(self):
        """Récupérer les relevés créés avec filtres"""
        queryset = ReleveDepense.objects.select_related(
            'valide_par', 'depenses_validees_par', 'cheque__banque', 'cheque__cree_par'
        ).prefetch_related('demandes')
        
        # Filtre par recherche (numéro de relevé)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(numero__icontains=search) |
                Q(observation__icontains=search)
            )
        
        # Filtre par période (date de création)
        date_debut = self.request.GET.get('date_debut')
        date_fin = self.request.GET.get('date_fin')
        if date_debut:
            try:
                from datetime import datetime
                date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
                queryset = queryset.filter(date_creation__date__gte=date_debut_obj)
            except (ValueError, TypeError):
                pass
        if date_fin:
            try:
                from datetime import datetime
                date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
                queryset = queryset.filter(date_creation__date__lte=date_fin_obj)
            except (ValueError, TypeError):
                pass
        
        # Filtre par année et mois (basé sur la période du relevé)
        annee = self.request.GET.get('annee')
        mois = self.request.GET.get('mois')
        if annee:
            try:
                annee_int = int(annee)
                queryset = queryset.filter(periode__year=annee_int)
                if mois:
                    try:
                        mois_int = int(mois)
                        queryset = queryset.filter(periode__month=mois_int)
                    except (ValueError, TypeError):
                        pass
            except (ValueError, TypeError):
                pass
        
        # Filtre par validé par (utilisateur)
        valide_par = self.request.GET.get('valide_par')
        if valide_par:
            try:
                queryset = queryset.filter(valide_par_id=valide_par)
            except (ValueError, TypeError):
                pass
        
        # Filtre par devise (montant > 0)
        devise = self.request.GET.get('devise')
        if devise == 'CDF':
            queryset = queryset.filter(montant_cdf__gt=0)
        elif devise == 'USD':
            queryset = queryset.filter(montant_usd__gt=0)
        
        return queryset.order_by('-date_creation')
    
    def get_context_data(self, **kwargs):
        """Ajouter les données de contexte pour les filtres"""
        context = super().get_context_data(**kwargs)
        
        # Récupérer les années disponibles (basées sur periode)
        annees = ReleveDepense.objects.exclude(periode__isnull=True).values_list('periode__year', flat=True).distinct().order_by('-periode__year')
        context['annees'] = list(annees)
        
        # Liste des utilisateurs ayant validé des relevés
        from accounts.models import User
        context['validateurs'] = User.objects.filter(
            releves_depense_valides__isnull=False
        ).distinct().order_by('username')
        
        # Paramètres de filtrage actuels
        context['filtres'] = {
            'search': self.request.GET.get('search', ''),
            'date_debut': self.request.GET.get('date_debut', ''),
            'date_fin': self.request.GET.get('date_fin', ''),
            'annee': self.request.GET.get('annee', ''),
            'mois': self.request.GET.get('mois', ''),
            'valide_par': self.request.GET.get('valide_par', ''),
            'devise': self.request.GET.get('devise', ''),
        }
        
        return context
    
    def dispatch(self, request, *args, **kwargs):
        """Gérer la requête avec gestion d'erreur pour éviter les problèmes de session"""
        try:
            return super().dispatch(request, *args, **kwargs)
        except Exception as e:
            # Logger l'erreur pour le débogage
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erreur dans ReleveDepenseListCreatedView: {str(e)}", exc_info=True)
            messages.error(request, 'Une erreur est survenue lors du chargement de la liste des relevés.')
            return redirect('demandes:releves_liste')


class ReleveDepenseDetailView(LoginRequiredMixin, DetailView):
    model = ReleveDepense
    template_name = 'demandes/releve_detail.html'
    context_object_name = 'releve'


class ChequePDFView(LoginRequiredMixin, View):
    """Vue pour générer un PDF de chèque pour un relevé de dépense"""
    
    def get(self, request, *args, **kwargs):
        releve_id = request.GET.get('releve_id')
        banque_id = request.GET.get('banque_id')
        
        if not releve_id:
            messages.error(request, 'Veuillez fournir un identifiant de relevé.')
            return redirect('demandes:releves_liste')
        
        if not banque_id:
            # Afficher le modal de sélection de banque
            from banques.models import Banque
            try:
                releve = ReleveDepense.objects.get(pk=releve_id)
            except ReleveDepense.DoesNotExist:
                messages.error(request, 'Relevé introuvable.')
                return redirect('demandes:releves_liste')
            
            banques = Banque.objects.filter(active=True).order_by('nom_banque')
            return render(request, 'demandes/cheque_banque_modal.html', {
                'releve': releve,
                'banques': banques
            })
        
        # Récupérer le relevé et la banque
        try:
            releve = ReleveDepense.objects.get(pk=releve_id)
            from banques.models import Banque
            banque = Banque.objects.get(pk=banque_id, active=True)
        except (ReleveDepense.DoesNotExist, Banque.DoesNotExist):
            messages.error(request, 'Relevé ou banque introuvable.')
            return redirect('demandes:releves_liste')
        
        # Créer ou récupérer le chèque
        # Utiliser les montants totaux des demandes (montant_cdf et montant_usd) au lieu du net à payer
        with transaction.atomic():
            cheque, created = Cheque.objects.get_or_create(
                releve_depense=releve,
                defaults={
                    'banque': banque,
                    'montant_cdf': releve.montant_cdf,  # Montant total CDF (avant IPR)
                    'montant_usd': releve.montant_usd,  # Montant total USD (avant IPR)
                    'cree_par': request.user,
                    'statut': 'GENERE'
                }
            )
            
            # Si le chèque existe déjà, mettre à jour la banque et les montants si nécessaire
            if not created:
                updated = False
                if cheque.banque != banque:
                    cheque.banque = banque
                    updated = True
                # Mettre à jour les montants avec les montants totaux (au lieu du net à payer)
                if cheque.montant_cdf != releve.montant_cdf or cheque.montant_usd != releve.montant_usd:
                    cheque.montant_cdf = releve.montant_cdf
                    cheque.montant_usd = releve.montant_usd
                    updated = True
                if updated:
                    cheque.save()
        
        # Créer le PDF du chèque
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="cheque_{cheque.numero_cheque}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        # Créer le document PDF
        doc = SimpleDocTemplate(response, pagesize=A4, 
                                rightMargin=2*cm, leftMargin=2*cm, 
                                topMargin=2*cm, bottomMargin=2*cm)
        elements = []
        
        # Styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'ChequeTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a3a5f'),
            alignment=TA_CENTER,
            spaceAfter=30,
        )
        
        normal_style = ParagraphStyle(
            'Normal',
            parent=styles['Normal'],
            fontSize=12,
            alignment=TA_LEFT,
        )
        
        # Titre
        title = Paragraph("CHÈQUE", title_style)
        elements.append(title)
        elements.append(Spacer(1, 1*cm))
        
        # Informations du chèque
        cheque_data = [
            ['Numéro de chèque:', cheque.numero_cheque],
            ['Banque:', banque.nom_banque],
            ['Relevé de dépense:', releve.numero],
            ['Date:', timezone.now().strftime('%d/%m/%Y')],
        ]
        
        if banque.adresse:
            cheque_data.append(['Adresse:', banque.adresse])
        
        # Tableau des informations
        info_table = Table(cheque_data, colWidths=[5*cm, 10*cm])
        info_table.setStyle(TableStyle([
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 11),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
        ]))
        elements.append(info_table)
        elements.append(Spacer(1, 1*cm))
        
        # Fonction pour convertir en lettres (copie de la fonction du template)
        def nombre_en_lettres(nombre):
            """Convertit un nombre en lettres en français"""
            if nombre is None or nombre == 0:
                return "zéro francs"
            
            try:
                from decimal import Decimal
                nombre = Decimal(str(nombre))
                nombre = abs(nombre)
                
                partie_entiere = int(nombre)
                partie_decimale = int(round((nombre - partie_entiere) * 100))
                
                unites = ['', 'un', 'deux', 'trois', 'quatre', 'cinq', 'six', 'sept', 'huit', 'neuf']
                dizaines_simples = ['', '', 'vingt', 'trente', 'quarante', 'cinquante', 'soixante']
                
                def convertir_deux_chiffres(n):
                    if n == 0:
                        return ''
                    if n < 10:
                        return unites[n]
                    if n < 20:
                        exceptions = {
                            10: 'dix', 11: 'onze', 12: 'douze', 13: 'treize', 14: 'quatorze',
                            15: 'quinze', 16: 'seize', 17: 'dix-sept', 18: 'dix-huit', 19: 'dix-neuf'
                        }
                        return exceptions.get(n, '')
                    
                    dizaine = n // 10
                    unite = n % 10
                    
                    if dizaine == 7:
                        if unite == 0:
                            return 'soixante-dix'
                        elif unite == 1:
                            return 'soixante-et-onze'
                        else:
                            return 'soixante-' + convertir_deux_chiffres(10 + unite)
                    elif dizaine == 8:
                        if unite == 0:
                            return 'quatre-vingts'
                        else:
                            return 'quatre-vingt-' + unites[unite]
                    elif dizaine == 9:
                        if unite == 0:
                            return 'quatre-vingt-dix'
                        else:
                            return 'quatre-vingt-' + convertir_deux_chiffres(10 + unite)
                    else:
                        if unite == 0:
                            return dizaines_simples[dizaine]
                        elif unite == 1 and dizaine > 1:
                            return dizaines_simples[dizaine] + '-et-un'
                        else:
                            return dizaines_simples[dizaine] + '-' + unites[unite]
                
                def convertir_trois_chiffres(n):
                    if n == 0:
                        return ''
                    
                    centaines = n // 100
                    reste = n % 100
                    
                    resultat = []
                    if centaines > 0:
                        if centaines == 1:
                            resultat.append('cent')
                        else:
                            resultat.append(unites[centaines] + ' cent')
                        if reste == 0 and centaines > 1:
                            resultat[-1] += 's'
                    
                    if reste > 0:
                        resultat.append(convertir_deux_chiffres(reste))
                    
                    return ' '.join(resultat)
                
                if partie_entiere == 0:
                    texte_entier = 'zéro'
                else:
                    groupes = []
                    millions = partie_entiere // 1000000
                    reste_millions = partie_entiere % 1000000
                    milliers = reste_millions // 1000
                    reste_milliers = reste_millions % 1000
                    
                    if millions > 0:
                        texte_millions = convertir_trois_chiffres(millions)
                        if millions == 1:
                            groupes.append('un million')
                        else:
                            groupes.append(texte_millions + ' millions')
                    
                    if milliers > 0:
                        texte_milliers = convertir_trois_chiffres(milliers)
                        if milliers == 1:
                            groupes.append('mille')
                        else:
                            groupes.append(texte_milliers + ' mille')
                    
                    if reste_milliers > 0:
                        groupes.append(convertir_trois_chiffres(reste_milliers))
                    
                    texte_entier = ' '.join(groupes)
                
                texte_decimal = ''
                if partie_decimale > 0:
                    texte_decimal = convertir_trois_chiffres(partie_decimale)
                    if partie_decimale == 1:
                        texte_decimal += ' centime'
                    else:
                        texte_decimal += ' centimes'
                
                if texte_decimal:
                    return f"{texte_entier} francs et {texte_decimal}"
                else:
                    if partie_entiere == 1:
                        return f"{texte_entier} franc"
                    else:
                        return f"{texte_entier} francs"
            
            except (ValueError, TypeError):
                return str(nombre) if nombre else "zéro francs"
        
        # Montants en chiffres et en lettres
        montant_cdf_chiffres = f"{cheque.montant_cdf:,.2f}".replace(',', ' ').replace('.', ',') + ' CDF'
        montant_cdf_lettres = nombre_en_lettres(cheque.montant_cdf)
        montant_usd_chiffres = f"{cheque.montant_usd:,.2f}".replace(',', ' ').replace('.', ',') + ' USD'
        montant_usd_lettres = nombre_en_lettres(cheque.montant_usd)
        
        # Style pour les montants en lettres
        lettres_style = ParagraphStyle(
            'LettresStyle',
            parent=styles['Normal'],
            fontSize=9,
            fontName='Helvetica-Oblique',
            textColor=colors.HexColor('#666666'),
            alignment=TA_LEFT,
        )
        
        # Tableau des montants
        montant_data = [
            ['Désignation', 'Montant'],
            ['Montant CDF', Paragraph(f"{montant_cdf_chiffres}<br/><i>{montant_cdf_lettres}</i>", normal_style)],
            ['Montant USD', Paragraph(f"{montant_usd_chiffres}<br/><i>{montant_usd_lettres}</i>", normal_style)],
        ]
        
        montant_table = Table(montant_data, colWidths=[8*cm, 7*cm])
        montant_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1a3a5f')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('FONTSIZE', (0, 1), (0, -1), 11),
            ('FONTSIZE', (1, 1), (1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        elements.append(montant_table)
        
        # Construire le PDF
        doc.build(elements)
        return response


class BanqueSoldesView(LoginRequiredMixin, View):
    """Vue AJAX pour récupérer les soldes des comptes bancaires d'une banque et comparer avec les montants du relevé"""
    
    def get(self, request, *args, **kwargs):
        banque_id = request.GET.get('banque_id')
        releve_id = request.GET.get('releve_id')
        
        if not banque_id:
            return JsonResponse({'error': 'banque_id requis'}, status=400)
        
        try:
            banque = Banque.objects.get(pk=banque_id, active=True)
        except Banque.DoesNotExist:
            return JsonResponse({'error': 'Banque introuvable'}, status=404)
        
        # Récupérer les montants totaux du relevé si fourni (utiliser montant_cdf et montant_usd au lieu de net_a_payer)
        montant_cdf_releve = None
        montant_usd_releve = None
        if releve_id:
            try:
                releve = ReleveDepense.objects.get(pk=releve_id)
                # Utiliser les montants totaux des demandes (avant IPR) au lieu du net à payer
                montant_cdf_releve = float(releve.montant_cdf) if releve.montant_cdf else 0.0
                montant_usd_releve = float(releve.montant_usd) if releve.montant_usd else 0.0
            except ReleveDepense.DoesNotExist:
                pass
        
        # Récupérer les comptes actifs de la banque
        comptes = CompteBancaire.objects.filter(banque=banque, actif=True).order_by('devise')
        
        soldes = {
            'banque': banque.nom_banque,
            'comptes': {},
            'avertissements': {}
        }
        
        for compte in comptes:
            solde = float(compte.solde_courant)
            montant_releve = montant_cdf_releve if compte.devise == 'CDF' else montant_usd_releve
            
            soldes['comptes'][compte.devise] = {
                'solde': solde,
                'intitule': compte.intitule_compte,
                'numero_compte': compte.numero_compte
            }
            
            # Vérifier si le montant du relevé dépasse le solde (pour les deux monnaies)
            # Si montant_releve est None, utiliser 0 pour la vérification
            montant_a_verifier = montant_releve if montant_releve is not None else 0.0
            
            # Vérifier que le montant est inférieur ou égal au solde
            if montant_a_verifier > 0:
                if solde < montant_a_verifier:
                    ecart = montant_a_verifier - solde
                    soldes['avertissements'][compte.devise] = {
                        'montant_releve': montant_a_verifier,
                        'solde': solde,
                        'ecart': ecart
                    }
        
        # Vérifier que les deux montants (USD et CDF) sont vérifiés
        # Si un montant est supérieur à 0 mais qu'il n'y a pas de compte correspondant, créer un avertissement
        if montant_cdf_releve is not None and montant_cdf_releve > 0:
            if 'CDF' not in soldes['comptes']:
                soldes['avertissements']['CDF'] = {
                    'montant_releve': montant_cdf_releve,
                    'solde': 0.0,
                    'ecart': montant_cdf_releve,
                    'message': 'Aucun compte CDF actif trouvé pour cette banque'
                }
        
        if montant_usd_releve is not None and montant_usd_releve > 0:
            if 'USD' not in soldes['comptes']:
                soldes['avertissements']['USD'] = {
                    'montant_releve': montant_usd_releve,
                    'solde': 0.0,
                    'ecart': montant_usd_releve,
                    'message': 'Aucun compte USD actif trouvé pour cette banque'
                }
        
        return JsonResponse(soldes)


class ChequeListView(LoginRequiredMixin, ListView):
    """Vue pour afficher la liste de tous les chèques"""
    model = Cheque
    template_name = 'demandes/cheque_liste.html'
    context_object_name = 'cheques'
    paginate_by = 20
    
    def get_queryset(self):
        """Récupérer tous les chèques avec optimisations"""
        queryset = Cheque.objects.select_related(
            'releve_depense', 'banque', 'cree_par'
        ).order_by('-date_creation')
        
        # Filtre par recherche (numéro de chèque, numéro de relevé)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(numero_cheque__icontains=search) |
                Q(releve_depense__numero__icontains=search) |
                Q(beneficiaire__icontains=search)
            )
        
        # Filtre par statut
        statut = self.request.GET.get('statut')
        if statut:
            queryset = queryset.filter(statut=statut)
        
        # Filtre par banque
        banque_id = self.request.GET.get('banque')
        if banque_id:
            queryset = queryset.filter(banque_id=banque_id)
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        from banques.models import Banque
        
        context['banques'] = Banque.objects.filter(active=True).order_by('nom_banque')
        context['statuts'] = Cheque.STATUT_CHOICES
        context['filtres'] = {
            'search': self.request.GET.get('search', ''),
            'statut': self.request.GET.get('statut', ''),
            'banque': self.request.GET.get('banque', ''),
        }
        return context


class ChequeDetailView(LoginRequiredMixin, DetailView):
    """Vue pour afficher les détails d'un chèque"""
    model = Cheque
    template_name = 'demandes/cheque_detail.html'
    context_object_name = 'cheque'
    
    def get_queryset(self):
        return Cheque.objects.select_related(
            'releve_depense', 'banque', 'cree_par'
        )


class ReleveDepenseValiderDepensesView(LoginRequiredMixin, View):
    """Vue pour valider les dépenses d'un relevé (créer des objets Depense à partir des demandes)"""
    
    def post(self, request, pk, *args, **kwargs):
        releve = get_object_or_404(ReleveDepense, pk=pk)
        
        # Vérifier si les dépenses sont déjà validées
        if releve.depenses_validees:
            messages.warning(request, 'Les dépenses de ce relevé ont déjà été validées.')
            return redirect('demandes:releve_detail', pk=releve.pk)
        
        # Vérifier les permissions
        if not request.user.peut_valider_depense():
            messages.error(request, 'Vous n\'avez pas la permission de valider les dépenses.')
            return redirect('demandes:releve_detail', pk=releve.pk)
        
        # Récupérer les demandes du relevé
        demandes = releve.demandes.all()
        
        if not demandes.exists():
            messages.error(request, 'Ce relevé ne contient aucune demande.')
            return redirect('demandes:releve_detail', pk=releve.pk)
        
        # Créer les objets Depense à partir des demandes
        depenses_creees = []
        from datetime import datetime
        
        for demande in demandes:
            # Générer le code de dépense
            date_depense = demande.date_demande or demande.date_soumission.date()
            annee = date_depense.year
            mois = date_depense.month
            
            # Chercher le dernier code pour ce mois/année
            from django.db.models import Max as MaxFunc
            last_code = Depense.objects.filter(
                code_depense__startswith=f'DEP-{annee}-{mois:02d}-'
            ).aggregate(MaxFunc('code_depense'))['code_depense__max']
            
            if last_code:
                try:
                    last_num = int(last_code.split('-')[-1])
                    new_num = last_num + 1
                except (ValueError, IndexError):
                    new_num = 1
            else:
                new_num = 1
            
            code_depense = f'DEP-{annee}-{mois:02d}-{new_num:04d}'
            
            # Vérifier si une dépense existe déjà pour cette demande
            depense_existante = Depense.objects.filter(
                code_depense=code_depense
            ).first()
            
            if depense_existante:
                continue  # Passer à la demande suivante si la dépense existe déjà
            
            # Créer l'objet Depense
            depense = Depense.objects.create(
                code_depense=code_depense,
                mois=mois,
                annee=annee,
                date_depense=date_depense,
                date_demande=demande.date_demande,
                nomenclature=demande.nomenclature,
                libelle_depenses=demande.description,
                montant_fc=demande.montant if demande.devise == 'CDF' else Decimal('0.00'),
                montant_usd=demande.montant if demande.devise == 'USD' else Decimal('0.00'),
                observation=f'Dépense validée depuis le relevé {releve.numero} - Demande {demande.reference}'
            )
            
            depenses_creees.append(depense)
        
        # Marquer le relevé comme ayant ses dépenses validées
        releve.depenses_validees = True
        releve.depenses_validees_par = request.user
        releve.date_validation_depenses = timezone.now()
        releve.save()
        
        messages.success(
            request, 
            f'{len(depenses_creees)} dépense(s) créée(s) avec succès à partir du relevé {releve.numero}.'
        )
        
        return redirect('demandes:releve_detail', pk=releve.pk)


class ReleveDepenseOldListView(LoginRequiredMixin, ListView):
    """Vue pour afficher les demandes validées comme relevé de dépense (ancienne version)"""
    model = DemandePaiement
    template_name = 'demandes/releve_liste_simple.html'
    context_object_name = 'demandes'
    paginate_by = 20
    
    def get_queryset(self):
        """Récupérer uniquement les demandes validées qui ne sont pas déjà dans un relevé"""
        queryset = DemandePaiement.objects.select_related(
            'service_demandeur', 'cree_par', 'approuve_par', 'nature_economique'
        ).filter(
            statut__in=['VALIDEE_DG', 'VALIDEE_DF', 'PAYEE']
        ).exclude(
            releves_depense__isnull=False  # Exclure les demandes déjà dans un relevé
        )
        
        # Filtrage selon le rôle
        if not self.request.user.peut_consulter_tout():
            if self.request.user.is_chef_service:
                queryset = queryset.filter(service_demandeur=self.request.user.service)
        
        # Trier par code de nature économique, puis par date
        return queryset.order_by('nature_economique__code', '-date_soumission')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculer les totaux pour toutes les demandes validées
        queryset = self.get_queryset()
        
        # Montants par devise
        montant_cdf = queryset.filter(devise='CDF').aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
        montant_usd = queryset.filter(devise='USD').aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
        
        # IPR (3%)
        ipr_cdf = montant_cdf * Decimal('0.03')
        ipr_usd = montant_usd * Decimal('0.03')
        
        # Net à payer (montant - IPR)
        net_a_payer_cdf = montant_cdf - ipr_cdf
        net_a_payer_usd = montant_usd - ipr_usd
        
        # Total général
        total_general = net_a_payer_cdf + net_a_payer_usd
        
        context['montant_cdf'] = montant_cdf
        context['montant_usd'] = montant_usd
        context['ipr_cdf'] = ipr_cdf
        context['ipr_usd'] = ipr_usd
        context['net_a_payer_cdf'] = net_a_payer_cdf
        context['net_a_payer_usd'] = net_a_payer_usd
        context['total_general'] = total_general
        
        # Grouper les demandes par code de nature économique pour les sous-totaux
        # et ajouter un numéro d'ordre à chaque demande
        demandes_groupes = {}
        demande_numero = {}
        numero = 1
        
        for demande in queryset:
            code = demande.nature_economique.code if demande.nature_economique else 'Sans code'
            if code not in demandes_groupes:
                demandes_groupes[code] = []
            demandes_groupes[code].append(demande)
            demande_numero[demande.pk] = numero
            numero += 1
        
        # Calculer les sous-totaux pour chaque groupe
        sous_totaux = {}
        for code, demandes_groupe in demandes_groupes.items():
            montant_usd_groupe = sum(d.montant for d in demandes_groupe if d.devise == 'USD')
            montant_cdf_groupe = sum(d.montant for d in demandes_groupe if d.devise == 'CDF')
            ipr_usd_groupe = montant_usd_groupe * Decimal('0.03')
            ipr_cdf_groupe = montant_cdf_groupe * Decimal('0.03')
            net_usd_groupe = montant_usd_groupe - ipr_usd_groupe
            net_cdf_groupe = montant_cdf_groupe - ipr_cdf_groupe
            
            sous_totaux[code] = {
                'montant_usd': montant_usd_groupe,
                'montant_cdf': montant_cdf_groupe,
                'ipr_usd': ipr_usd_groupe,
                'ipr_cdf': ipr_cdf_groupe,
                'net_usd': net_usd_groupe,
                'net_cdf': net_cdf_groupe,
            }
        
        context['sous_totaux'] = sous_totaux
        context['demande_numero'] = demande_numero
        
        return context


class ReleveDepenseCreateView(LoginRequiredMixin, View):
    """Vue pour créer automatiquement un relevé avec toutes les demandes validées"""
    
    def post(self, request, *args, **kwargs):
        # Vérifier les permissions (seuls DAF et DG peuvent créer des relevés)
        if not request.user.peut_valider_depense():
            messages.error(request, 'Vous n\'avez pas la permission de créer un relevé de dépense.')
            return redirect('demandes:releves_liste_old')
        
        # Récupérer toutes les demandes validées non encore dans un relevé
        demandes_valides = DemandePaiement.objects.select_related(
            'service_demandeur', 'cree_par', 'approuve_par', 'nature_economique'
        ).filter(
            statut__in=['VALIDEE_DG', 'VALIDEE_DF', 'PAYEE']
        ).exclude(
            releves_depense__isnull=False
        )
        
        if not demandes_valides.exists():
            messages.warning(request, 'Aucune demande validée disponible pour créer un relevé.')
            return redirect('demandes:releves_liste_old')
        
        try:
            with transaction.atomic():
                # Créer le relevé avec une période unique
                from datetime import date, datetime, timedelta
                now = timezone.now()
                
                # Utiliser le premier jour du mois comme période, mais avec une heure unique
                periode_unique = now.date().replace(day=1)
                
                # Vérifier si un relevé existe déjà pour cette période
                existing_count = ReleveDepense.objects.filter(periode=periode_unique).count()
                if existing_count > 0:
                    # Ajouter un delta de temps pour rendre la période unique
                    periode_unique = periode_unique + timedelta(days=existing_count)
                
                numero = f"REL-{now.year}-{now.month:02d}-{ReleveDepense.objects.count() + 1:06d}"
                
                releve = ReleveDepense.objects.create(
                    numero=numero,
                    periode=periode_unique,
                    valide_par=request.user,
                    date_validation=now
                )
                
                # Ajouter toutes les demandes au relevé
                for demande in demandes_valides:
                    releve.demandes.add(demande)
                
                # Calculer les totaux du relevé
                releve.calculer_total()
                
                messages.success(
                    request, 
                    f'Relevé {releve.numero} créé avec succès contenant {len(demandes_valides)} demande(s).'
                )
                
                return redirect('demandes:releves_liste_old')
                
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du relevé : {str(e)}')
            return redirect('demandes:releves_liste_old')
    
    def get(self, request, *args, **kwargs):
        # Pour les requêtes GET, rediriger vers la liste des demandes
        return redirect('demandes:releves_liste_old')


class ReleveDepenseAjouterDemandesView(LoginRequiredMixin, UpdateView):
    """Vue pour ajouter des demandes à un relevé existant"""
    model = ReleveDepense
    form_class = ReleveDepenseForm
    template_name = 'demandes/releve_ajouter_demandes.html'
    success_url = reverse_lazy('demandes:releves_liste')
    
    def dispatch(self, request, *args, **kwargs):
        # Vérifier les permissions
        if not request.user.peut_valider_depense():
            messages.error(request, 'Vous n\'avez pas la permission de modifier un relevé de dépense.')
            return redirect('demandes:releves_liste')
        return super().dispatch(request, *args, **kwargs)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        releve = self.get_object()
        
        # Ajouter le relevé au contexte avec le nom attendu par le template
        context['releve'] = releve
        
        # Récupérer les demandes validées non encore dans un relevé
        demandes_disponibles = DemandePaiement.objects.filter(
            statut__in=['VALIDEE_DG', 'VALIDEE_DF', 'PAYEE']
        ).exclude(
            releves_depense__isnull=False
        ).select_related('service_demandeur', 'nature_economique')
        
        context['demandes_disponibles'] = demandes_disponibles
        context['demandes_actuelles'] = releve.demandes.all()
        
        return context
    
    def post(self, request, *args, **kwargs):
        releve = self.get_object()
        
        # Récupérer les demandes sélectionnées depuis le formulaire
        demandes_selectionnees = request.POST.getlist('demandes')
        
        if demandes_selectionnees:
            # Créer un queryset à partir des IDs
            demandes_queryset = DemandePaiement.objects.filter(
                pk__in=demandes_selectionnees
            )
            
            # Utiliser la méthode sécurisée pour ajouter les demandes
            try:
                demandes_ajoutees = releve.ajouter_demandes_securise(demandes_queryset)
                releve.calculer_total()
                
                messages.success(
                    request, 
                    f'{len(demandes_ajoutees)} demande(s) ajoutée(s) au relevé avec succès.'
                )
            except Exception as e:
                messages.error(request, f'Erreur lors de l\'ajout des demandes : {str(e)}')
                return redirect('demandes:releve_ajouter_demandes', pk=releve.pk)
        else:
            messages.info(request, 'Aucune demande sélectionnée.')
        
        return redirect('demandes:releve_detail', pk=releve.pk)
    
    def form_valid(self, form):
        # Cette méthode n'est plus utilisée, nous utilisons post() directement
        return redirect('demandes:releve_detail', pk=self.get_object().pk)


class ReleveDepenseAncienneCreateView(LoginRequiredMixin, CreateView):
    """Ancienne vue pour créer un relevé avec demandes (gardée pour compatibilité)"""
    model = ReleveDepense
    form_class = ReleveDepenseForm
    template_name = 'demandes/releve_form.html'
    success_url = reverse_lazy('demandes:releves_liste')
    
    def form_valid(self, form):
        releve = form.save(commit=False)
        releve.valide_par = self.request.user
        releve.save()
        
        # Vérifier que seules les demandes validées sont ajoutées
        demandes_selectionnees = form.cleaned_data.get('demandes', [])
        if demandes_selectionnees:
            # Créer un queryset à partir de la liste
            from django.db.models import Q
            demandes_queryset = DemandePaiement.objects.filter(
                pk__in=[d.pk for d in demandes_selectionnees]
            )
            
            # Utiliser la méthode sécurisée pour éviter les doublons
            try:
                releve.ajouter_demandes_securise(demandes_queryset)
            except Exception as e:
                messages.error(self.request, f'Erreur lors de l\'ajout des demandes : {str(e)}')
                return redirect('demandes:releve_detail', pk=releve.pk)
        else:
            form.save_m2m()
        
        releve.calculer_total()
        messages.success(self.request, 'Relevé de dépense créé avec succès.')
        return redirect(self.success_url)


class ReleveDepenseAutoCreateView(LoginRequiredMixin, FormView):
    """Vue pour générer automatiquement un relevé à partir des demandes validées"""
    form_class = ReleveDepenseAutoForm
    template_name = 'demandes/releve_auto_form.html'
    success_url = reverse_lazy('demandes:releves_liste')
    
    def dispatch(self, request, *args, **kwargs):
        # Vérifier les permissions (seuls DAF et DG peuvent générer des relevés)
        if not request.user.peut_valider_depense():
            messages.error(request, 'Vous n\'avez pas la permission de générer un relevé de dépense.')
            return redirect('demandes:releves_liste')
        return super().dispatch(request, *args, **kwargs)
    
    def form_valid(self, form):
        periode = form.cleaned_data['periode']
        observation = form.cleaned_data.get('observation', '')
        
        # Vérifier si un relevé existe déjà pour cette période
        releve_existant = ReleveDepense.objects.filter(periode=periode).first()
        if releve_existant:
            messages.warning(
                self.request, 
                f'Un relevé existe déjà pour la période {periode.strftime("%d/%m/%Y")}. '
                f'Voulez-vous le modifier ou en créer un nouveau ?'
            )
            return redirect('demandes:releves_liste')
        
        # Filtrer les demandes validées (VALIDEE_DG ou VALIDEE_DF) pour cette période
        # Filtrer par date_demande si disponible, sinon par date_soumission
        annee = periode.year
        mois = periode.month
        
        # Récupérer toutes les demandes validées (toutes devises)
        demandes_validees = DemandePaiement.objects.filter(
            statut__in=['VALIDEE_DG', 'VALIDEE_DF', 'PAYEE']
        )
        
        # Filtrer par période : utiliser date_demande si disponible, sinon date_soumission
        demandes_validees = demandes_validees.filter(
            Q(date_demande__year=annee, date_demande__month=mois) |
            Q(date_demande__isnull=True, date_soumission__year=annee, date_soumission__month=mois)
        )
        
        # Exclure les demandes déjà dans un relevé
        demandes_validees = demandes_validees.exclude(releves_depense__isnull=False)
        
        if not demandes_validees.exists():
            messages.warning(
                self.request, 
                f'Aucune demande validée trouvée pour la période {periode.strftime("%d/%m/%Y")}.'
            )
            return self.form_invalid(form)
        
        # Créer le relevé
        releve = ReleveDepense.objects.create(
            periode=periode,
            observation=observation,
            valide_par=self.request.user
        )
        
        # Ajouter toutes les demandes validées de manière sécurisée
        try:
            demandes_ajoutees = releve.ajouter_demandes_securise(demandes_validees)
            releve.calculer_total()
        except Exception as e:
            messages.error(self.request, f'Erreur lors de l\'ajout des demandes : {str(e)}')
            releve.delete()  # Supprimer le relevé créé
            return self.form_invalid(form)
        
        messages.success(
            self.request, 
            f'Relevé de dépense généré avec succès ! {demandes_validees.count()} demande(s) incluse(s). '
            f'Total général : {releve.get_total_general()}'
        )
        
        return redirect(self.success_url)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Statistiques des demandes validées disponibles (toutes devises)
        demandes_validees = DemandePaiement.objects.filter(
            statut__in=['VALIDEE_DG', 'VALIDEE_DF', 'PAYEE']
        ).exclude(releves_depense__isnull=False)
        
        demandes_validees_usd = demandes_validees.filter(devise='USD').count()
        demandes_validees_cdf = demandes_validees.filter(devise='CDF').count()
        
        context['demandes_validees_usd'] = demandes_validees_usd
        context['demandes_validees_cdf'] = demandes_validees_cdf
        
        return context


class DepenseListView(LoginRequiredMixin, ListView):
    """Vue pour lister les demandes validées comme dépenses"""
    model = DemandePaiement
    template_name = 'demandes/depense_liste.html'
    context_object_name = 'depenses'
    paginate_by = 50
    
    def get_queryset(self):
        # Récupérer les demandes qui sont dans des relevés dont les dépenses ont été validées
        queryset = DemandePaiement.objects.select_related(
            'service_demandeur', 'cree_par', 'approuve_par', 'nature_economique', 'nomenclature'
        ).prefetch_related('releves_depense').filter(
            releves_depense__depenses_validees=True
        ).distinct()
        
        # Filtrage par relevé (numéro de relevé)
        releve_numero = self.request.GET.get('releve')
        if releve_numero:
            queryset = queryset.filter(releves_depense__numero=releve_numero)
        
        # Filtrage par plage de dates (sur date_demande ou date_soumission si date_demande est None)
        date_debut = self.request.GET.get('date_debut')
        date_fin = self.request.GET.get('date_fin')
        if date_debut:
            try:
                from datetime import datetime
                date_debut_obj = datetime.strptime(date_debut, '%Y-%m-%d').date()
                # Filtrer sur date_demande si elle existe, sinon sur date_soumission
                queryset = queryset.filter(
                    Q(date_demande__gte=date_debut_obj) | 
                    Q(date_demande__isnull=True, date_soumission__date__gte=date_debut_obj)
                )
            except (ValueError, TypeError):
                pass
        if date_fin:
            try:
                from datetime import datetime
                date_fin_obj = datetime.strptime(date_fin, '%Y-%m-%d').date()
                # Filtrer sur date_demande si elle existe, sinon sur date_soumission
                queryset = queryset.filter(
                    Q(date_demande__lte=date_fin_obj) | 
                    Q(date_demande__isnull=True, date_soumission__date__lte=date_fin_obj)
                )
            except (ValueError, TypeError):
                pass
        
        # Filtrage par année (basé sur date_demande) - seulement si pas de filtre par date
        if not date_debut and not date_fin:
            annee = self.request.GET.get('annee')
            if annee:
                try:
                    queryset = queryset.filter(date_demande__year=int(annee))
                except (ValueError, TypeError):
                    pass
            
            # Filtrage par mois (basé sur date_demande)
            mois = self.request.GET.get('mois')
            if mois:
                try:
                    queryset = queryset.filter(date_demande__month=int(mois))
                except (ValueError, TypeError):
                    pass
        
        # Filtrage par nature économique
        nature_economique_id = self.request.GET.get('nature_economique')
        if nature_economique_id:
            queryset = queryset.filter(nature_economique_id=nature_economique_id)
        
        # Filtrage par service
        service_id = self.request.GET.get('service')
        if service_id:
            queryset = queryset.filter(service_demandeur_id=service_id)
        
        # Filtrage par devise
        devise = self.request.GET.get('devise')
        if devise:
            queryset = queryset.filter(devise=devise)
        
        # Recherche textuelle
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(description__icontains=search) |
                Q(reference__icontains=search)
            )
        
        return queryset.order_by('-date_demande', '-date_soumission')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Calculer les totaux
        queryset = self.get_queryset()
        total_cdf = sum(d.montant for d in queryset.filter(devise='CDF'))
        total_usd = sum(d.montant for d in queryset.filter(devise='USD'))
        context['total_fc'] = total_cdf
        context['total_usd'] = total_usd
        
        # Liste des années disponibles (basées sur date_demande)
        annees = DemandePaiement.objects.filter(
            releves_depense__depenses_validees=True
        ).exclude(date_demande__isnull=True).values_list('date_demande__year', flat=True).distinct().order_by('-date_demande__year')
        context['annees'] = list(annees)
        
        # Liste des natures économiques
        context['natures_economiques'] = NatureEconomique.objects.all().order_by('code')
        
        # Liste des services
        from accounts.models import Service
        context['services'] = Service.objects.filter(actif=True).order_by('nom_service')
        
        # Récupérer le relevé actuel si filtré
        releve_numero = self.request.GET.get('releve')
        releve_actuel = None
        if releve_numero:
            try:
                releve_actuel = ReleveDepense.objects.select_related('valide_par', 'depenses_validees_par').get(
                    numero=releve_numero,
                    depenses_validees=True
                )
            except ReleveDepense.DoesNotExist:
                pass
        
        context['releve_actuel'] = releve_actuel
        
        # Paramètres de filtrage actuels
        context['filtres'] = {
            'date_debut': self.request.GET.get('date_debut', ''),
            'date_fin': self.request.GET.get('date_fin', ''),
            'annee': self.request.GET.get('annee', ''),
            'mois': self.request.GET.get('mois', ''),
            'nature_economique': self.request.GET.get('nature_economique', ''),
            'service': self.request.GET.get('service', ''),
            'devise': self.request.GET.get('devise', ''),
            'search': self.request.GET.get('search', ''),
            'releve': releve_numero or '',
        }
        
        return context


class DepenseCreateView(LoginRequiredMixin, CreateView):
    """Vue pour créer une nouvelle dépense"""
    model = Depense
    form_class = DepenseForm
    template_name = 'demandes/depense_form.html'
    success_url = reverse_lazy('demandes:depenses_liste')
    
    def form_valid(self, form):
        messages.success(self.request, 'Dépense créée avec succès.')
        return super().form_valid(form)


class DepenseUpdateView(LoginRequiredMixin, UpdateView):
    """Vue pour modifier une dépense existante"""
    model = Depense
    form_class = DepenseForm
    template_name = 'demandes/depense_form.html'
    success_url = reverse_lazy('demandes:depenses_liste')
    
    def form_valid(self, form):
        messages.success(self.request, 'Dépense modifiée avec succès.')
        return super().form_valid(form)


class DepenseDetailView(LoginRequiredMixin, DetailView):
    """Vue pour afficher les détails d'une dépense"""
    model = Depense
    template_name = 'demandes/depense_detail.html'
    context_object_name = 'depense'


def get_nomenclatures_by_year(request):
    """API pour récupérer les nomenclatures par année"""
    from .models import NomenclatureDepense
    annee = request.GET.get('annee')
    
    if annee:
        try:
            annee = int(annee)
            nomenclatures = NomenclatureDepense.objects.filter(
                statut='EN_COURS',
                annee=annee
            ).order_by('-date_publication')
            
            data = []
            for n in nomenclatures:
                data.append({
                    'id': n.id,
                    'libelle_article': f"Nomenclature {n.annee} - {n.date_publication}",
                    'code_compte': str(n.annee)
                })
            
            return JsonResponse(data, safe=False)
        except ValueError:
            return JsonResponse({'error': 'Année invalide'}, status=400)
    
    return JsonResponse({'error': 'Année requise'}, status=400)


class NatureEconomiqueListView(LoginRequiredMixin, ListView):
    """Vue pour lister les natures économiques"""
    model = NatureEconomique
    template_name = 'demandes/nature_liste.html'
    context_object_name = 'natures'
    paginate_by = 50
    
    def get_queryset(self):
        queryset = NatureEconomique.objects.select_related('parent').all()
        
        # Recherche textuelle
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(
                Q(code__icontains=search) |
                Q(titre__icontains=search) |
                Q(description__icontains=search)
            )
        
        # Filtrage par parent
        parent_id = self.request.GET.get('parent')
        if parent_id:
            queryset = queryset.filter(parent_id=parent_id)
        
        return queryset.order_by('code')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Liste des natures parentes (racines)
        context['parents'] = NatureEconomique.objects.filter(parent__isnull=True).order_by('code')
        
        # Paramètres de filtrage actuels
        context['filtres'] = {
            'parent': self.request.GET.get('parent', ''),
            'search': self.request.GET.get('search', ''),
        }
        
        return context


class NatureEconomiqueCreateView(LoginRequiredMixin, CreateView):
    """Vue pour créer une nouvelle nature économique"""
    model = NatureEconomique
    form_class = NatureEconomiqueForm
    template_name = 'demandes/nature_form.html'
    success_url = reverse_lazy('demandes:nature_liste')
    
    def form_valid(self, form):
        messages.success(self.request, 'Nature économique créée avec succès.')
        return super().form_valid(form)


class NatureEconomiqueUpdateView(LoginRequiredMixin, UpdateView):
    """Vue pour modifier une nature économique existante"""
    model = NatureEconomique
    form_class = NatureEconomiqueForm
    template_name = 'demandes/nature_form.html'
    success_url = reverse_lazy('demandes:nature_liste')
    
    def form_valid(self, form):
        messages.success(self.request, 'Nature économique modifiée avec succès.')
        return super().form_valid(form)


class NatureEconomiqueDetailView(LoginRequiredMixin, DetailView):
    """Vue pour afficher les détails d'une nature économique"""
    model = NatureEconomique
    template_name = 'demandes/nature_detail.html'
    context_object_name = 'nature'


# ==================== VUES DE PAIEMENT ====================

class PaiementCreateView(LoginRequiredMixin, CreateView):
    """Vue pour créer un paiement pour une demande"""
    model = Paiement
    form_class = PaiementForm
    template_name = 'demandes/paiement_form.html'
    success_url = reverse_lazy('demandes:paiement_liste')
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        
        # Passer le paramètre GET 'releve' au formulaire
        releve_id = self.request.GET.get('releve')
        if releve_id:
            kwargs['initial'] = kwargs.get('initial', {})
            kwargs['initial']['releve_depense'] = releve_id
            
            # Pré-remplir le queryset des demandes et mettre à jour les montants
            try:
                releve = ReleveDepense.objects.get(pk=releve_id)
                demandes = releve.demandes.all()
                
                # Mettre à jour les montants pour les demandes qui ne sont pas payées
                for demande in demandes:
                    if demande.reste_a_payer == 0 and demande.statut != 'PAYEE':
                        demande.reste_a_payer = demande.montant - demande.montant_deja_paye
                        demande.save()
                
                kwargs['releve_queryset'] = demandes
            except ReleveDepense.DoesNotExist:
                kwargs['releve_queryset'] = DemandePaiement.objects.none()
        else:
            kwargs['releve_queryset'] = DemandePaiement.objects.none()
        
        return kwargs
    
    def form_valid(self, form):
        form.instance.paiement_par = self.request.user
        messages.success(self.request, 'Paiement effectué avec succès!')
        return super().form_valid(form)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Effectuer un paiement'
        return context


class PaiementListView(LoginRequiredMixin, ListView):
    """Vue pour lister les paiements"""
    model = Paiement
    template_name = 'demandes/paiement_liste.html'
    context_object_name = 'paiements'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = Paiement.objects.select_related(
            'demande', 'releve_depense', 'paiement_par'
        ).prefetch_related('demande__service_demandeur')
        
        # Filtrage par relevé de dépenses
        releve_id = self.request.GET.get('releve')

        if releve_id:
            queryset = queryset.filter(releve_depense_id=releve_id)
        
        # Filtrage par devise
        devise = self.request.GET.get('devise')
        if devise:
            queryset = queryset.filter(devise=devise)
        
        return queryset.order_by('-date_paiement')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Historique des paiements'
        
        # Récupérer les relevés pour le filtre
        context['releves'] = self.get_queryset().distinct()

        #context['releves'] = self.get_queryset().values("releve_depense_id")
        return context


class PaiementParReleveView(LoginRequiredMixin, FormView):
    """Vue pour payer les demandes d'un relevé de dépenses"""
    template_name = 'demandes/paiement_releve.html'
    form_class = PaiementMultipleForm
    
    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs['user'] = self.request.user
        return kwargs
    
    def get_form(self, form_class=None):
        return super().get_form(form_class)
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'Paiement par relevé de dépenses'
        return context
    
    def form_valid(self, form):
        releve_depense = form.cleaned_data['releve_depense']
        return redirect('demandes:paiement_releve_detail', pk=releve_depense.pk)


class PaiementReleveDetailView(LoginRequiredMixin, View):
    """Vue pour afficher et payer les demandes d'un relevé de dépenses spécifique"""
    
    def get(self, request, pk):
        releve_depense = get_object_or_404(
            ReleveDepense.objects,
            pk=pk
        )
        
        # Récupérer les demandes de ce relevé (toutes, pas seulement non payées)
        toutes_demandes = releve_depense.demandes.all().select_related('service_demandeur')
        
        # Forcer le recalcul des montants pour chaque demande
        for demande in toutes_demandes:
            # S'assurer que les valeurs Decimal sont valides
            try:
                if demande.montant is None:
                    demande.montant = Decimal('0.00')
                if demande.montant_deja_paye is None:
                    demande.montant_deja_paye = Decimal('0.00')
                if demande.reste_a_payer is None:
                    demande.reste_a_payer = demande.montant - demande.montant_deja_paye
                
                # Convertir en Decimal si ce n'est pas déjà le cas
                demande.montant = Decimal(str(demande.montant))
                demande.montant_deja_paye = Decimal(str(demande.montant_deja_paye))
                demande.reste_a_payer = Decimal(str(demande.reste_a_payer))
                
                # Recalculer si nécessaire
                if demande.reste_a_payer == 0 and demande.statut != 'PAYEE':
                    demande.reste_a_payer = demande.montant - demande.montant_deja_paye
                    demande.save()
            except (ValueError, TypeError, Exception) as e:
                # Si erreur, recalculer depuis les valeurs de base
                try:
                    demande.montant = Decimal(str(demande.montant)) if demande.montant else Decimal('0.00')
                    demande.montant_deja_paye = Decimal(str(demande.montant_deja_paye)) if demande.montant_deja_paye else Decimal('0.00')
                    demande.reste_a_payer = demande.montant - demande.montant_deja_paye
                    demande.save()
                except Exception:
                    # Ignorer cette demande si elle cause toujours des problèmes
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(f"Impossible de traiter la demande {demande.pk}: {e}")
                    continue
        
        # Maintenant filtrer les demandes non entièrement payées
        demandes = toutes_demandes.filter(
            reste_a_payer__gt=0
        )
        
        # Calculer les totaux en gérant les valeurs None ou invalides
        total_a_payer = Decimal('0.00')
        total_initial = Decimal('0.00')
        total_deja_paye = Decimal('0.00')
        
        for demande in toutes_demandes:
            # Total initial
            if demande.montant is not None:
                try:
                    total_initial += Decimal(str(demande.montant))
                except (ValueError, TypeError, Exception):
                    pass
            
            # Total déjà payé
            if demande.montant_deja_paye is not None:
                try:
                    total_deja_paye += Decimal(str(demande.montant_deja_paye))
                except (ValueError, TypeError, Exception):
                    pass
        
        # Total à payer (seulement pour les demandes non entièrement payées)
        for demande in demandes:
            if demande.reste_a_payer is not None:
                try:
                    total_a_payer += Decimal(str(demande.reste_a_payer))
                except (ValueError, TypeError, Exception):
                    pass
        
        # Déterminer la devise du relevé (utiliser la devise de la première demande si disponible)
        devise_releve = 'CDF'  # Par défaut
        if toutes_demandes.exists():
            premiere_demande = toutes_demandes.first()
            if premiere_demande and premiere_demande.devise:
                devise_releve = premiere_demande.devise
        
        # Calculer le total du relevé (montant_cdf + montant_usd converti ou utiliser total_initial)
        total_releve = Decimal('0.00')
        if releve_depense.montant_cdf:
            try:
                total_releve += Decimal(str(releve_depense.montant_cdf))
            except (ValueError, TypeError, Exception):
                pass
        if releve_depense.montant_usd:
            try:
                total_releve += Decimal(str(releve_depense.montant_usd))
            except (ValueError, TypeError, Exception):
                pass
        
        # Si total_releve est 0, utiliser total_initial
        if total_releve == 0:
            total_releve = total_initial
        
        # Construire le titre avec le numéro du relevé
        titre_releve = f"Relevé {releve_depense.numero}" if releve_depense.numero else f"Relevé du {releve_depense.periode}"
        
        # Récupérer les comptes bancaires actifs pour la sélection
        from banques.models import CompteBancaire
        comptes_bancaires = CompteBancaire.objects.filter(actif=True).select_related('banque').order_by('banque__nom_banque', 'devise')
        
        context = {
            'title': f'Paiement des demandes du {titre_releve}',
            'releve_depense': releve_depense,
            'demandes': demandes,
            'toutes_demandes': toutes_demandes,  # Ajouter toutes les demandes pour le résumé
            'total_demandes': demandes.count(),
            'total_a_payer': total_a_payer,
            'total_initial': total_initial,
            'total_deja_paye': total_deja_paye,
            'total_releve': total_releve,
            'devise_releve': devise_releve,
            'comptes_bancaires': comptes_bancaires,
        }
        
        return render(request, 'demandes/paiement_releve_detail.html', context)
    
    def payer_demande(self, request, pk, demande_pk):
        """Méthode pour payer une seule demande"""
        releve_depense = get_object_or_404(ReleveDepense.objects, pk=pk)
        demande = get_object_or_404(DemandePaiement.objects, pk=demande_pk)
        
        # Vérifier que la demande appartient au relevé
        if demande not in releve_depense.demandes.all():
            messages.error(request, "Cette demande n'appartient pas à ce relevé.")
            return redirect('demandes:paiement_releve_detail', pk=pk)
        
        # Récupérer les données du formulaire
        montant_key = f'montant_{demande.pk}'
        beneficiaire_key = f'beneficiaire_{demande.pk}'
        observations_key = f'observations_{demande.pk}'
        
        try:
            # Nettoyer la valeur avant conversion
            montant_str = request.POST.get(montant_key, '0.00').strip().replace(',', '.').replace(' ', '')
            if not montant_str:
                montant_str = '0.00'
            
            montant = Decimal(montant_str)
            beneficiaire = request.POST.get(beneficiaire_key, '').strip()
            observations = request.POST.get(observations_key, '').strip()
            
            if montant <= 0:
                messages.error(request, f"Le montant doit être supérieur à 0 pour la demande {demande.reference}.")
                return redirect('demandes:paiement_releve_detail', pk=pk)
            
            if not beneficiaire:
                messages.error(request, f"Veuillez spécifier le bénéficiaire pour la demande {demande.reference}.")
                return redirect('demandes:paiement_releve_detail', pk=pk)
            
            # Vérifier que le montant ne dépasse pas le reste à payer
            if montant > demande.reste_a_payer:
                messages.error(request, f"Le montant ({montant}) ne peut pas dépasser le reste à payer ({demande.reste_a_payer}) pour la demande {demande.reference}.")
                return redirect('demandes:paiement_releve_detail', pk=pk)
            
            # Le compte bancaire n'est plus requis pour le paiement
            # Créer le paiement
            with transaction.atomic():
                paiement = Paiement.objects.create(
                    demande=demande,
                    releve_depense=releve_depense,
                    paiement_par=request.user,
                    montant_paye=montant,
                    devise=demande.devise,
                    beneficiaire=beneficiaire,
                    observations=observations or f"Paiement pour la demande {demande.reference}"
                )
                
                # Le compte bancaire n'est plus utilisé pour le paiement
                # La mise à jour du solde bancaire doit être faite manuellement si nécessaire
                
                # Mettre à jour la demande
                demande.montant_deja_paye += montant
                
                # S'assurer que reste_a_payer est un Decimal valide avant l'opération
                if demande.reste_a_payer is None:
                    demande.reste_a_payer = demande.montant - demande.montant_deja_paye
                else:
                    try:
                        demande.reste_a_payer = Decimal(str(demande.reste_a_payer)) - montant
                    except (ValueError, TypeError, Exception):
                        demande.reste_a_payer = demande.montant - demande.montant_deja_paye
                
                if demande.reste_a_payer <= 0:
                    demande.statut = 'PAYEE'
                    demande.reste_a_payer = Decimal('0.00')
                
                demande.save()
            
            messages.success(request, f"Paiement de {montant} {demande.devise} enregistré avec succès pour la demande {demande.reference}.")
            
        except (ValueError, TypeError, Exception) as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erreur lors du paiement de la demande {demande.pk}: {e}")
            messages.error(request, f"Une erreur est survenue lors de l'enregistrement du paiement: {str(e)}")
        
        return redirect('demandes:paiement_releve_detail', pk=pk)
    
    def post(self, request, pk):
        # Vérifier si c'est un paiement individuel
        demande_pk = request.POST.get('demande_pk')
        if demande_pk:
            return self.payer_demande(request, pk, demande_pk)
        
        # Sinon, traitement du paiement multiple
        releve_depense = get_object_or_404(
            ReleveDepense.objects,
            pk=pk
        )
        
        # Récupérer les montants à payer pour chaque demande
        paiements_data = {}
        total_paye = Decimal('0.00')
        
        for demande in releve_depense.demandes.all():
            montant_key = f'montant_{demande.pk}'
            beneficiaire_key = f'beneficiaire_{demande.pk}'
            observations_key = f'observations_{demande.pk}'
            
            if montant_key in request.POST:
                try:
                    # Nettoyer la valeur avant conversion (enlever espaces, remplacer virgule par point)
                    montant_str = request.POST.get(montant_key, '0.00').strip().replace(',', '.').replace(' ', '')
                    if not montant_str:
                        montant_str = '0.00'
                    
                    montant = Decimal(montant_str)
                    beneficiaire = request.POST.get(beneficiaire_key, '').strip()
                    observations = request.POST.get(observations_key, '').strip()
                    
                    if montant > 0 and beneficiaire:  # Exiger le bénéficiaire
                        # Le compte bancaire n'est plus requis pour le paiement
                        paiements_data[demande.pk] = {
                            'montant': montant,
                            'demande': demande,
                            'beneficiaire': beneficiaire,
                            'observations': observations,
                            'compte_bancaire': None  # Plus utilisé
                        }
                        total_paye += montant
                    elif montant > 0 and not beneficiaire:
                        messages.error(request, f"Veuillez spécifier le bénéficiaire pour la demande {demande.reference}.")
                        return self.get(request, pk)
                except (ValueError, TypeError, Exception) as e:
                    # Logger l'erreur pour le débogage
                    import logging
                    logger = logging.getLogger(__name__)
                    logger.warning(f"Erreur de conversion Decimal pour {montant_key}: {e}, valeur: {request.POST.get(montant_key, 'N/A')}")
                    continue
        
        if not paiements_data:
            messages.error(request, "Veuillez saisir au moins un montant à payer.")
            return self.get(request, pk)
        
        # Créer les paiements
        with transaction.atomic():
            for demande_pk, paiement_data in paiements_data.items():
                demande = paiement_data['demande']
                montant = paiement_data['montant']
                beneficiaire = paiement_data['beneficiaire']
                observations = paiement_data['observations']
                compte_bancaire = paiement_data.get('compte_bancaire')
                
                # Créer le paiement
                # Rafraîchir le solde une dernière fois avant la déduction pour éviter les problèmes de concurrence
                if compte_bancaire:
                    compte_bancaire.refresh_from_db()
                    solde_avant_paiement = compte_bancaire.solde_courant
                    
                    # Vérifier à nouveau le solde dans la transaction
                    if solde_avant_paiement < montant:
                        messages.error(request, f"Solde insuffisant. Solde disponible: {solde_avant_paiement} {compte_bancaire.devise}, Montant requis: {montant} {demande.devise} pour la demande {demande.reference}.")
                        return self.get(request, pk)
                
                paiement = Paiement.objects.create(
                    demande=demande,
                    releve_depense=releve_depense,  # Associer au relevé de dépenses
                    paiement_par=request.user,
                    montant_paye=montant,
                    devise=demande.devise,
                    beneficiaire=beneficiaire,
                    observations=observations or f"Paiement par rapport au relevé {releve_depense.periode}"
                )
                
                # Le compte bancaire n'est plus utilisé pour le paiement
                # La mise à jour du solde bancaire doit être faite manuellement si nécessaire
                
                # Mettre à jour la demande
                demande.montant_deja_paye += montant
                
                # S'assurer que reste_a_payer est un Decimal valide avant l'opération
                if demande.reste_a_payer is None:
                    demande.reste_a_payer = demande.montant - demande.montant_deja_paye
                else:
                    try:
                        demande.reste_a_payer = Decimal(str(demande.reste_a_payer)) - montant
                    except (ValueError, TypeError, Exception):
                        demande.reste_a_payer = demande.montant - demande.montant_deja_paye
                
                if demande.reste_a_payer <= 0:
                    demande.statut = 'PAYEE'
                    demande.reste_a_payer = Decimal('0.00')
                
                demande.save()
        
        messages.success(request, f"{len(paiements_data)} paiement(s) effectué(s) avec succès pour un total de {total_paye} {paiements_data[next(iter(paiements_data))]['demande'].devise}.")
        return redirect('demandes:paiement_releve_detail', pk=pk)
    
    def payer_demande(self, request, pk, demande_pk):
        """Méthode pour payer une seule demande"""
        releve_depense = get_object_or_404(ReleveDepense.objects, pk=pk)
        demande = get_object_or_404(DemandePaiement.objects, pk=demande_pk)
        
        # Vérifier que la demande appartient au relevé
        if demande not in releve_depense.demandes.all():
            messages.error(request, "Cette demande n'appartient pas à ce relevé.")
            return redirect('demandes:paiement_releve_detail', pk=pk)
        
        # Récupérer les données du formulaire
        montant_key = f'montant_{demande.pk}'
        beneficiaire_key = f'beneficiaire_{demande.pk}'
        observations_key = f'observations_{demande.pk}'
        
        try:
            # Nettoyer la valeur avant conversion
            montant_str = request.POST.get(montant_key, '0.00').strip().replace(',', '.').replace(' ', '')
            if not montant_str:
                montant_str = '0.00'
            
            montant = Decimal(montant_str)
            beneficiaire = request.POST.get(beneficiaire_key, '').strip()
            observations = request.POST.get(observations_key, '').strip()
            
            if montant <= 0:
                messages.error(request, f"Le montant doit être supérieur à 0 pour la demande {demande.reference}.")
                return redirect('demandes:paiement_releve_detail', pk=pk)
            
            if not beneficiaire:
                messages.error(request, f"Veuillez spécifier le bénéficiaire pour la demande {demande.reference}.")
                return redirect('demandes:paiement_releve_detail', pk=pk)
            
            # Vérifier que le montant ne dépasse pas le reste à payer
            if montant > demande.reste_a_payer:
                messages.error(request, f"Le montant ({montant}) ne peut pas dépasser le reste à payer ({demande.reste_a_payer}) pour la demande {demande.reference}.")
                return redirect('demandes:paiement_releve_detail', pk=pk)
            
            # Le compte bancaire n'est plus requis pour le paiement
            # Créer le paiement
            with transaction.atomic():
                paiement = Paiement.objects.create(
                    demande=demande,
                    releve_depense=releve_depense,
                    paiement_par=request.user,
                    montant_paye=montant,
                    devise=demande.devise,
                    beneficiaire=beneficiaire,
                    observations=observations or f"Paiement pour la demande {demande.reference}"
                )
                
                # Le compte bancaire n'est plus utilisé pour le paiement
                # La mise à jour du solde bancaire doit être faite manuellement si nécessaire
                
                # Mettre à jour la demande
                demande.montant_deja_paye += montant
                
                # S'assurer que reste_a_payer est un Decimal valide avant l'opération
                if demande.reste_a_payer is None:
                    demande.reste_a_payer = demande.montant - demande.montant_deja_paye
                else:
                    try:
                        demande.reste_a_payer = Decimal(str(demande.reste_a_payer)) - montant
                    except (ValueError, TypeError, Exception):
                        demande.reste_a_payer = demande.montant - demande.montant_deja_paye
                
                if demande.reste_a_payer <= 0:
                    demande.statut = 'PAYEE'
                    demande.reste_a_payer = Decimal('0.00')
                
                demande.save()
            
            messages.success(request, f"Paiement de {montant} {demande.devise} enregistré avec succès pour la demande {demande.reference}.")
            
        except (ValueError, TypeError, Exception) as e:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(f"Erreur lors du paiement de la demande {demande.pk}: {e}")
            messages.error(request, f"Une erreur est survenue lors de l'enregistrement du paiement: {str(e)}")
        
        return redirect('demandes:paiement_releve_detail', pk=pk)


class PaiementDetailView(LoginRequiredMixin, DetailView):
    """Vue pour afficher les détails d'un paiement"""
    model = Paiement
    template_name = 'demandes/paiement_detail.html'
    context_object_name = 'paiement'
    
    def get_queryset(self):
        return Paiement.objects.select_related(
            'demande', 'releve_depense', 'paiement_par'
        ).prefetch_related('demande__service_demandeur')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = f'Détails du paiement {self.object.reference}'
        
        # Calculer le pourcentage de paiement
        if self.object.demande.montant > 0:
            pourcentage = (self.object.demande.montant_deja_paye / self.object.demande.montant) * 100
        else:
            pourcentage = 0
        context['pourcentage_paiement'] = pourcentage
        
        return context


# ==================== VUES API ====================

class DemandeResteAPayerView(LoginRequiredMixin, View):
    """Vue API pour obtenir le reste à payer d'une demande"""
    
    def get(self, request, pk):
        try:
            demande = DemandePaiement.objects.get(pk=pk)
            data = {
                'success': True,
                'montant_deja_paye': str(demande.montant_deja_paye),
                'reste_a_payer': str(demande.reste_a_payer),
                'montant_total': str(demande.montant),
                'devise': demande.devise
            }
            return JsonResponse(data)
        except DemandePaiement.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Demande non trouvée'
            }, status=404)

