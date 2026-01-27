"""
Vues pour la gestion des états
"""
import logging
from django.views.generic import ListView, CreateView, DetailView, View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.urls import reverse_lazy
from django.shortcuts import redirect, get_object_or_404, render
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.db.models import Q, Sum
from django.utils import timezone
from decimal import Decimal
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.platypus.doctemplate import PageTemplate, BaseDocTemplate
from reportlab.platypus.frames import Frame
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
import json

logger = logging.getLogger('etats')

from .models import EtatGenerique, ConfigurationEtat, HistoriqueGeneration
from .forms import EtatSelectionForm, FiltresAvancesForm
from demandes.models import DemandePaiement, ReleveDepense, Depense, Paiement, NatureEconomique
from recettes.models import Recette
from releves.models import ReleveBancaire
from accounts.models import Service
from banques.models import Banque, CompteBancaire


class EtatListView(LoginRequiredMixin, ListView):
    """Liste de tous les états générés"""
    model = EtatGenerique
    template_name = 'etats/etat_liste.html'
    context_object_name = 'etats'
    paginate_by = 20
    
    def get_queryset(self):
        queryset = EtatGenerique.objects.select_related('genere_par').prefetch_related(
            'services', 'natures_economiques', 'banques', 'comptes_bancaires'
        )
        
        # Filtrage par type
        type_etat = self.request.GET.get('type_etat')
        if type_etat:
            queryset = queryset.filter(type_etat=type_etat)
        
        # Filtrage par statut
        statut = self.request.GET.get('statut')
        if statut:
            queryset = queryset.filter(statut=statut)
        
        return queryset.order_by('-date_generation')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['types_etat'] = EtatGenerique.TYPE_ETAT_CHOICES
        context['statuts'] = EtatGenerique.STATUT_CHOICES
        return context


class EtatCreateView(LoginRequiredMixin, View):
    """Vue pour créer un nouvel état"""
    template_name = 'etats/etat_selection.html'
    
    def get(self, request, *args, **kwargs):
        form = EtatSelectionForm()
        form_filtres = FiltresAvancesForm()
        configurations = ConfigurationEtat.objects.filter(actif=True)
        
        return render(request, self.template_name, {
            'form': form,
            'form_filtres': form_filtres,
            'configurations': configurations,
        })
    
    def post(self, request, *args, **kwargs):
        form = EtatSelectionForm(request.POST)
        form_filtres = FiltresAvancesForm(request.POST)
        
        if form.is_valid() and form_filtres.is_valid():
            # Appliquer la périodicité si nécessaire
            if form.cleaned_data.get('periodicite') != 'PERSONNALISE':
                form.appliquer_periodicite()
            
            # Créer l'état
            etat = EtatGenerique.objects.create(
                titre=form.cleaned_data['titre'],
                type_etat=form.cleaned_data['type_etat'],
                description=form.cleaned_data['description'],
                date_debut=form.cleaned_data['date_debut'],
                date_fin=form.cleaned_data['date_fin'],
                periodicite=form.cleaned_data['periodicite'],
                genere_par=request.user,
                statut='GENERATION',
                filtres_supplementaires=form_filtres.cleaned_data,
                parametres_affichage={
                    'format_sortie': form.cleaned_data['format_sortie'],
                    'inclure_details': form.cleaned_data['inclure_details'],
                    'inclure_graphiques': form.cleaned_data['inclure_graphiques'],
                    'tri_par': form.cleaned_data['tri_par'],
                    'ordre_tri': form.cleaned_data['ordre_tri'],
                }
            )
            
            # Ajouter les relations many-to-many
            if form.cleaned_data['services']:
                etat.services.set(form.cleaned_data['services'])
            if form.cleaned_data['natures_economiques']:
                etat.natures_economiques.set(form.cleaned_data['natures_economiques'])
            if form.cleaned_data['banques']:
                etat.banques.set(form.cleaned_data['banques'])
            if form.cleaned_data['comptes_bancaires']:
                etat.comptes_bancaires.set(form.cleaned_data['comptes_bancaires'])
            
            messages.success(request, 'État créé avec succès. Génération en cours...')
            return redirect('etats:generer', pk=etat.pk)
        else:
            configurations = ConfigurationEtat.objects.filter(actif=True)
            return render(request, self.template_name, {
                'form': form,
                'form_filtres': form_filtres,
                'configurations': configurations,
            })


class EtatPreviewView(LoginRequiredMixin, View):
    """Vue pour prévisualiser les données avant génération"""
    
    def post(self, request):
        try:
            print("=== DEBUG PREVIEW START ===")
            
            # Récupérer les paramètres
            type_etat = request.POST.get('type_etat')
            date_debut_str = request.POST.get('date_debut')
            date_fin_str = request.POST.get('date_fin')
            services_ids = request.POST.getlist('services')
            natures_ids = request.POST.getlist('natures_economiques')
            banques_ids = request.POST.getlist('banques')
            comptes_ids = request.POST.getlist('comptes_bancaires')
            
            print(f"Type état: {type_etat}")
            print(f"Date début: {date_debut_str}")
            print(f"Date fin: {date_fin_str}")
            print(f"Services: {services_ids}")
            print(f"Natures: {natures_ids}")
            print(f"Banques: {banques_ids}")
            print(f"Comptes: {comptes_ids}")
            
            # Convertir les dates en objets date
            from datetime import datetime
            date_debut = None
            date_fin = None
            
            if date_debut_str:
                date_debut = datetime.strptime(date_debut_str, '%Y-%m-%d').date()
            if date_fin_str:
                date_fin = datetime.strptime(date_fin_str, '%Y-%m-%d').date()
            
            print(f"Date début: {date_debut} (type: {type(date_debut)})")
            print(f"Date fin: {date_fin} (type: {type(date_fin)})")
            
            # Calculer les données directement sans créer d'objet EtatGenerique
            donnees = self.calculer_donnees_direct(type_etat, date_debut, date_fin, 
                                                services_ids, natures_ids, banques_ids, comptes_ids)
            print(f"Données calculées: {donnees}")
            
            # Sérialiser les données pour le JSON de manière simple
            lignes_serialisees = []
            for ligne in donnees['lignes']:
                try:
                    if type_etat == 'RELEVE_DEPENSE':
                        ligne_data = {
                            'numero': str(ligne.numero) if ligne.numero else '',
                            'periode': ligne.periode.strftime('%d/%m/%Y') if ligne.periode else '',
                            'net_a_payer_usd': str(ligne.net_a_payer_usd),
                            'net_a_payer_cdf': str(ligne.net_a_payer_cdf),
                            'valide_par': str(ligne.valide_par) if ligne.valide_par else '',
                            'date_creation': ligne.date_creation.strftime('%d/%m/%Y') if ligne.date_creation else ''
                        }
                    elif type_etat == 'DEMANDE_PAIEMENT':
                        ligne_data = {
                            'reference': str(ligne.reference) if ligne.reference else '',
                            'service_demandeur': ligne.service_demandeur.nom_service if ligne.service_demandeur else '',
                            'nature_economique': str(ligne.nature_economique) if ligne.nature_economique else '',
                            'description': str(ligne.description) if ligne.description else '',
                            'montant': str(ligne.montant),
                            'devise': str(ligne.devise) if ligne.devise else '',
                            'statut': str(ligne.statut) if ligne.statut else '',
                            'date_soumission': ligne.date_soumission.strftime('%d/%m/%Y') if ligne.date_soumission else ''
                        }
                    elif type_etat == 'PAIEMENT':
                        ligne_data = {
                            'reference': str(ligne.reference) if ligne.reference else '',
                            'demande': str(ligne.demande) if ligne.demande else '',
                            'montant_paye': str(ligne.montant_paye),
                            'devise': str(ligne.devise) if ligne.devise else '',
                            'date_paiement': ligne.date_paiement.strftime('%d/%m/%Y') if ligne.date_paiement else '',
                            'paiement_par': str(ligne.paiement_par) if ligne.paiement_par else ''
                        }
                    else:
                        ligne_data = {'info': 'Type non géré'}
                    
                    lignes_serialisees.append(ligne_data)
                    print(f"Ligne sérialisée: {ligne_data}")
                    
                except Exception as e:
                    print(f"Erreur sérialisation ligne: {e}")
                    lignes_serialisees.append({'error': str(e)})
            
            response_data = {
                'success': True,
                'count': donnees['count'],
                'total_usd': float(donnees['total_usd']),
                'total_cdf': float(donnees['total_cdf']),
                'lignes': lignes_serialisees
            }
            
            print(f"Réponse JSON: {response_data}")
            print("=== DEBUG PREVIEW END ===")
            
            return JsonResponse(response_data)
            
        except Exception as e:
            print(f"Erreur générale dans preview: {e}")
            import traceback
            traceback.print_exc()
            return JsonResponse({
                'success': False,
                'error': str(e)
            })
    
    def calculer_donnees_direct(self, type_etat, date_debut, date_fin, services_ids, natures_ids, banques_ids, comptes_ids):
        """Calcule les données directement sans créer d'objet EtatGenerique"""
        if type_etat == 'DEMANDE_PAIEMENT':
            return self.calculer_donnees_demandes_direct(date_debut, date_fin, services_ids, natures_ids)
        elif type_etat == 'RECETTE':
            return self.calculer_donnees_recettes_direct(date_debut, date_fin, banques_ids, comptes_ids)
        elif type_etat == 'DEPENSE':
            return self.calculer_donnees_depenses_direct(date_debut, date_fin, banques_ids)
        elif type_etat == 'PAIEMENT':
            return self.calculer_donnees_paiements_direct(date_debut, date_fin)
        elif type_etat == 'RELEVE_DEPENSE':
            return self.calculer_donnees_releves_depenses_direct(date_debut, date_fin, banques_ids)
        elif type_etat == 'SOLDE_BANCAIRE':
            return self.calculer_donnees_soldes_direct(banques_ids, comptes_ids)
        else:
            return {
                'total_usd': Decimal('0.00'),
                'total_cdf': Decimal('0.00'),
                'lignes': [],
                'count': 0
            }
    
    def calculer_donnees_releves_depenses_direct(self, date_debut, date_fin, banques_ids):
        """Calcule les données pour les relevés de dépenses sans objet EtatGenerique"""
        queryset = ReleveDepense.objects.select_related('valide_par').prefetch_related('demandes')
        
        # Debug: Afficher les informations de base
        print(f"=== DEBUG RELEVÉS DIRECT ===")
        print(f"Période demandée: {date_debut} au {date_fin}")
        print(f"Banques: {banques_ids}")
        
        # Vérifier d'abord tous les relevés sans filtre
        tous_releves = ReleveDepense.objects.all()
        print(f"Total relevés dans la base: {tous_releves.count()}")
        
        # Afficher les périodes disponibles
        periodes_disponibles = tous_releves.values_list('periode', flat=True)
        print(f"Périodes disponibles: {list(periodes_disponibles)}")
        
        # CORRECTION: Pour les relevés de dépenses, la période est stockée comme premier jour du mois
        # Il faut donc inclure tous les relevés dont la période est entre date_debut et date_fin
        # en considérant que la période du relevé représente tout le mois
        
        # Si la période demandée est un seul jour, l'élargir au mois complet
        if date_debut == date_fin:
            # C'est une demande pour un jour spécifique, trouver le mois correspondant
            mois_debut = date_debut.replace(day=1)
            # Dernier jour du mois
            if date_debut.month == 12:
                mois_fin = date_debut.replace(date_debut.year + 1, 1, 1)
            else:
                mois_fin = date_debut.replace(date_debut.year, date_debut.month + 1, 1)
            mois_fin = mois_fin - timezone.timedelta(days=1)
            
            print(f"Période élargie au mois: {mois_debut} au {mois_fin}")
            queryset = queryset.filter(
                periode__gte=mois_debut,
                periode__lte=mois_fin
            )
        else:
            # Période normale, mais s'assurer d'inclure les mois complets
            queryset = queryset.filter(
                periode__gte=date_debut.replace(day=1),
                periode__lte=date_fin
            )
        
        print(f"Relevés après filtre période: {queryset.count()}")
        
        # Le modèle ReleveDepense n'a pas de champ banque, donc pas de filtrage par banque
        # if banques_ids:
        #     queryset = queryset.filter(banque__in=banques_ids)
        #     print(f"Relevés après filtre banques: {queryset.count()}")
        
        # Calculer les totaux
        total_usd = queryset.aggregate(total=Sum('net_a_payer_usd'))['total'] or Decimal('0.00')
        total_cdf = queryset.aggregate(total=Sum('net_a_payer_cdf'))['total'] or Decimal('0.00')
        
        print(f"Totaux calculés: USD={total_usd}, CDF={total_cdf}")
        print(f"========================")
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_demandes_direct(self, date_debut, date_fin, services_ids, natures_ids):
        """Calcule les données pour les demandes de paiement sans objet EtatGenerique"""
        queryset = DemandePaiement.objects.select_related('service_demandeur', 'nature_economique')
        
        # Appliquer les filtres de période
        queryset = queryset.filter(
            date_soumission__gte=date_debut,
            date_soumission__lte=date_fin
        )
        
        # Appliquer les filtres
        if services_ids:
            queryset = queryset.filter(service_demandeur__pk__in=services_ids)
        if natures_ids:
            queryset = queryset.filter(nature_economique__pk__in=natures_ids)
        
        # Calculer les totaux
        total_usd = queryset.filter(devise='USD').aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
        total_cdf = queryset.filter(devise='CDF').aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_recettes_direct(self, date_debut, date_fin, banques_ids, comptes_ids):
        """Calcule les données pour les recettes sans objet EtatGenerique"""
        queryset = Recette.objects.select_related('banque', 'compte_bancaire')
        
        # Appliquer les filtres de période
        queryset = queryset.filter(
            date_encaissement__gte=date_debut,
            date_encaissement__lte=date_fin
        )
        
        # Appliquer les filtres
        if banques_ids:
            queryset = queryset.filter(banque__pk__in=banques_ids)
        if comptes_ids:
            queryset = queryset.filter(compte_bancaire__pk__in=comptes_ids)
        
        # Calculer les totaux
        total_usd = queryset.aggregate(total=Sum('montant_usd'))['total'] or Decimal('0.00')
        total_cdf = queryset.aggregate(total=Sum('montant_cdf'))['total'] or Decimal('0.00')
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_depenses_direct(self, date_debut, date_fin, banques_ids):
        """Calcule les données pour les dépenses sans objet EtatGenerique"""
        queryset = Depense.objects.select_related('banque', 'nomenclature')
        
        # Appliquer les filtres de période
        queryset = queryset.filter(
            date_depense__gte=date_debut,
            date_depense__lte=date_fin
        )
        
        # Appliquer les filtres
        if banques_ids:
            queryset = queryset.filter(banque__pk__in=banques_ids)
        
        # Calculer les totaux - le modèle Depense a montant_usd et montant_fc, pas de champ devise
        total_usd = queryset.aggregate(total=Sum('montant_usd'))['total'] or Decimal('0.00')
        total_cdf = queryset.aggregate(total=Sum('montant_fc'))['total'] or Decimal('0.00')
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_paiements_direct(self, date_debut, date_fin):
        """Calcule les données pour les paiements sans objet EtatGenerique"""
        queryset = Paiement.objects.select_related('demande', 'demande__service_demandeur')
        
        # Appliquer les filtres de période
        queryset = queryset.filter(
            date_paiement__gte=date_debut,
            date_paiement__lte=date_fin
        )
        
        # Calculer les totaux
        total_usd = queryset.filter(devise='USD').aggregate(total=Sum('montant_paye'))['total'] or Decimal('0.00')
        total_cdf = queryset.filter(devise='CDF').aggregate(total=Sum('montant_paye'))['total'] or Decimal('0.00')
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_soldes_direct(self, banques_ids, comptes_ids):
        """Calcule les données pour les soldes bancaires sans objet EtatGenerique"""
        queryset = CompteBancaire.objects.select_related('banque').filter(actif=True)
        
        # Appliquer les filtres
        if banques_ids:
            queryset = queryset.filter(banque__pk__in=banques_ids)
        if comptes_ids:
            queryset = queryset.filter(pk__in=comptes_ids)
        
        # Calculer les totaux - le modèle CompteBancaire a un champ devise
        total_usd = queryset.filter(devise='USD').aggregate(total=Sum('solde_courant'))['total'] or Decimal('0.00')
        total_cdf = queryset.filter(devise='CDF').aggregate(total=Sum('solde_courant'))['total'] or Decimal('0.00')
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_par_type(self, etat):
        """Calcule les données selon le type d'état"""
        if etat.type_etat == 'DEMANDE_PAIEMENT':
            return self.calculer_donnees_demandes(etat)
        elif etat.type_etat == 'RECETTE':
            return self.calculer_donnees_recettes(etat)
        elif etat.type_etat == 'DEPENSE':
            return self.calculer_donnees_depenses(etat)
        elif etat.type_etat == 'PAIEMENT':
            return self.calculer_donnees_paiements(etat)
        elif etat.type_etat == 'RELEVE_DEPENSE':
            return self.calculer_donnees_releves_depenses(etat)
        elif etat.type_etat == 'SOLDE_BANCAIRE':
            return self.calculer_donnees_soldes_bancaires(etat)
        else:
            return {
                'total_usd': Decimal('0.00'),
                'total_cdf': Decimal('0.00'),
                'lignes': [],
                'count': 0
            }
    
    def calculer_donnees_demandes(self, etat):
        """Calcule les données pour les demandes de paiement"""
        queryset = DemandePaiement.objects.select_related('service_demandeur', 'nature_economique')
        
        # Appliquer les filtres de période
        queryset = queryset.filter(
            date_soumission__gte=etat.date_debut,
            date_soumission__lte=etat.date_fin
        )
        
        # Appliquer les filtres many-to-many
        if etat.services.exists():
            queryset = queryset.filter(service_demandeur__in=etat.services.all())
        if etat.natures_economiques.exists():
            queryset = queryset.filter(nature_economique__in=etat.natures_economiques.all())
        
        # Calculer les totaux
        total_usd = queryset.filter(devise='USD').aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
        total_cdf = queryset.filter(devise='CDF').aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_recettes(self, etat):
        """Calcule les données pour les recettes"""
        queryset = Recette.objects.select_related('banque', 'compte_bancaire')
        
        # Appliquer les filtres de période
        queryset = queryset.filter(
            date_encaissement__gte=etat.date_debut,
            date_encaissement__lte=etat.date_fin
        )
        
        # Appliquer les filtres many-to-many
        if etat.banques.exists():
            queryset = queryset.filter(banque__in=etat.banques.all())
        if etat.comptes_bancaires.exists():
            queryset = queryset.filter(compte_bancaire__in=etat.comptes_bancaires.all())
        
        # Calculer les totaux
        total_usd = queryset.aggregate(total=Sum('montant_usd'))['total'] or Decimal('0.00')
        total_cdf = queryset.aggregate(total=Sum('montant_cdf'))['total'] or Decimal('0.00')
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_depenses(self, etat):
        """Calcule les données pour les dépenses"""
        queryset = Depense.objects.select_related('banque', 'nomenclature')
        
        # Appliquer les filtres de période
        queryset = queryset.filter(
            date_depense__gte=etat.date_debut,
            date_depense__lte=etat.date_fin
        )
        
        # Appliquer les filtres many-to-many
        if etat.banques.exists():
            queryset = queryset.filter(banque__in=etat.banques.all())
        
        # Calculer les totaux - le modèle Depense a montant_usd et montant_fc, pas de champ devise
        total_usd = queryset.aggregate(total=Sum('montant_usd'))['total'] or Decimal('0.00')
        total_cdf = queryset.aggregate(total=Sum('montant_fc'))['total'] or Decimal('0.00')
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_paiements(self, etat):
        """Calcule les données pour les paiements"""
        queryset = Paiement.objects.select_related('demande', 'demande__service_demandeur')
        
        # Appliquer les filtres de période
        queryset = queryset.filter(
            date_paiement__gte=etat.date_debut,
            date_paiement__lte=etat.date_fin
        )
        
        # Calculer les totaux
        total_usd = queryset.filter(devise='USD').aggregate(total=Sum('montant_paye'))['total'] or Decimal('0.00')
        total_cdf = queryset.filter(devise='CDF').aggregate(total=Sum('montant_paye'))['total'] or Decimal('0.00')
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_releves_depenses(self, etat):
        """Calcule les données pour les relevés de dépenses"""
        queryset = ReleveDepense.objects.select_related('valide_par').prefetch_related('demandes')
        
        # Debug: Afficher les informations de base
        print(f"=== DEBUG RELEVÉS ===")
        print(f"Période demandée: {etat.date_debut} au {etat.date_fin}")
        print(f"Type état: {etat.type_etat}")
        
        # Vérifier d'abord tous les relevés sans filtre
        tous_releves = ReleveDepense.objects.all()
        print(f"Total relevés dans la base: {tous_releves.count()}")
        
        # Afficher les périodes disponibles
        periodes_disponibles = tous_releves.values_list('periode', flat=True)
        print(f"Périodes disponibles: {list(periodes_disponibles)}")
        
        # CORRECTION: Pour les relevés de dépenses, la période est stockée comme premier jour du mois
        # Il faut donc inclure tous les relevés dont la période est entre date_debut et date_fin
        # en considérant que la période du relevé représente tout le mois
        
        # Si la période demandée est un seul jour, l'élargir au mois complet
        if etat.date_debut == etat.date_fin:
            # C'est une demande pour un jour spécifique, trouver le mois correspondant
            mois_debut = etat.date_debut.replace(day=1)
            # Dernier jour du mois
            if etat.date_debut.month == 12:
                mois_fin = etat.date_debut.replace(etat.date_debut.year + 1, 1, 1)
            else:
                mois_fin = etat.date_debut.replace(etat.date_debut.year, etat.date_debut.month + 1, 1)
            mois_fin = mois_fin - timezone.timedelta(days=1)
            
            print(f"Période élargie au mois: {mois_debut} au {mois_fin}")
            queryset = queryset.filter(
                periode__gte=mois_debut,
                periode__lte=mois_fin
            )
        else:
            # Période normale, mais s'assurer d'inclure les mois complets
            queryset = queryset.filter(
                periode__gte=etat.date_debut.replace(day=1),
                periode__lte=etat.date_fin
            )
        
        print(f"Relevés après filtre période: {queryset.count()}")
        
        # Le modèle ReleveDepense n'a pas de champ banque, donc pas de filtrage par banque
        # if etat.banques.exists():
        #     queryset = queryset.filter(banque__in=etat.banques.all())
        #     print(f"Relevés après filtre banques: {queryset.count()}")
        
        # Calculer les totaux
        total_usd = queryset.aggregate(total=Sum('net_a_payer_usd'))['total'] or Decimal('0.00')
        total_cdf = queryset.aggregate(total=Sum('net_a_payer_cdf'))['total'] or Decimal('0.00')
        
        print(f"Totaux calculés: USD={total_usd}, CDF={total_cdf}")
        print(f"========================")
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_soldes_bancaires(self, etat):
        """Calcule les données pour les soldes bancaires"""
        queryset = CompteBancaire.objects.select_related('banque').filter(actif=True)
        
        # Appliquer les filtres many-to-many
        if etat.banques.exists():
            queryset = queryset.filter(banque__in=etat.banques.all())
        if etat.comptes_bancaires.exists():
            queryset = queryset.filter(pk__in=etat.comptes_bancaires.all())
        
        # Calculer les totaux - le modèle CompteBancaire a un champ devise
        total_usd = queryset.filter(devise='USD').aggregate(total=Sum('solde_courant'))['total'] or Decimal('0.00')
        total_cdf = queryset.filter(devise='CDF').aggregate(total=Sum('solde_courant'))['total'] or Decimal('0.00')
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }


@method_decorator(csrf_exempt, name='dispatch')
class EtatCreateAjaxView(LoginRequiredMixin, View):
    """Vue pour créer un état via AJAX"""
    
    def post(self, request):
        try:
            print("=== DEBUG CREATE AJAX START ===")
            print(f"POST data: {request.POST}")
            
            form = EtatSelectionForm(request.POST)
            form_filtres = FiltresAvancesForm(request.POST)
            
            print(f"Form valid: {form.is_valid()}")
            print(f"Form errors: {form.errors}")
            print(f"Filtres valid: {form_filtres.is_valid()}")
            print(f"Filtres errors: {form_filtres.errors}")
            
            if form.is_valid() and form_filtres.is_valid():
                print("Forms are valid, creating etat...")
                
                # Appliquer la périodicité si nécessaire
                if form.cleaned_data.get('periodicite') != 'PERSONNALISE':
                    form.appliquer_periodicite()
                
                # Créer l'état
                etat = EtatGenerique.objects.create(
                    titre=form.cleaned_data['titre'],
                    type_etat=form.cleaned_data['type_etat'],
                    description=form.cleaned_data['description'],
                    date_debut=form.cleaned_data['date_debut'],
                    date_fin=form.cleaned_data['date_fin'],
                    periodicite=form.cleaned_data['periodicite'],
                    genere_par=request.user,
                    statut='GENERATION',
                    filtres_supplementaires=form_filtres.cleaned_data,
                    parametres_affichage={
                        'format_sortie': form.cleaned_data['format_sortie'],
                        'inclure_details': form.cleaned_data['inclure_details'],
                        'inclure_graphiques': form.cleaned_data['inclure_graphiques'],
                        'tri_par': form.cleaned_data['tri_par'],
                        'ordre_tri': form.cleaned_data['ordre_tri'],
                    }
                )
                
                print(f"Etat created with ID: {etat.pk}")
                
                # Ajouter les relations many-to-many
                if form.cleaned_data['services']:
                    etat.services.set(form.cleaned_data['services'])
                if form.cleaned_data['natures_economiques']:
                    etat.natures_economiques.set(form.cleaned_data['natures_economiques'])
                if form.cleaned_data['banques']:
                    etat.banques.set(form.cleaned_data['banques'])
                if form.cleaned_data['comptes_bancaires']:
                    etat.comptes_bancaires.set(form.cleaned_data['comptes_bancaires'])
                
                print("=== DEBUG CREATE AJAX END ===")
                
                return JsonResponse({
                    'success': True,
                    'etat_id': etat.pk
                })
            else:
                print("=== DEBUG CREATE AJAX FAILED ===")
                return JsonResponse({
                    'success': False,
                    'error': 'Formulaire invalide',
                    'form_errors': form.errors,
                    'filtres_errors': form_filtres.errors
                })
                
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': str(e)
            })


class EtatDetailView(LoginRequiredMixin, DetailView):
    """Vue pour voir les détails d'un état"""
    model = EtatGenerique
    template_name = 'etats/etat_detail.html'
    context_object_name = 'etat'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        etat = self.get_object()
        context['historique'] = HistoriqueGeneration.objects.filter(
            etat=etat
        ).select_related('utilisateur').order_by('-date_action')
        return context


class EtatGenererView(LoginRequiredMixin, DetailView):
    """Vue pour générer les fichiers d'un état"""
    model = EtatGenerique
    template_name = 'etats/etat_generer.html'
    context_object_name = 'etat'
    
    def get(self, request, *args, **kwargs):
        etat = self.get_object()
        
        try:
            # Calculer les données selon le type d'état
            donnees = self.calculer_donnees(etat)
            
            # Mettre à jour les totaux
            etat.total_usd = donnees.get('total_usd', Decimal('0.00'))
            etat.total_cdf = donnees.get('total_cdf', Decimal('0.00'))
            etat.save()
            
            # Générer les fichiers selon le format demandé
            format_sortie = etat.parametres_affichage.get('format_sortie', 'PDF')
            
            if format_sortie in ['PDF', 'LES_DEUX']:
                self.generer_pdf(etat, donnees)
            
            if format_sortie in ['EXCEL', 'LES_DEUX']:
                self.generer_excel(etat, donnees)
            
            # Marquer comme généré
            etat.statut = 'GENERE'
            etat.save()
            
            # Message informatif selon le nombre de résultats
            if donnees.get('count', 0) == 0:
                messages.info(request, f'État généré avec succès! Aucune donnée trouvée pour la période sélectionnée.')
            else:
                messages.success(request, f'État généré avec succès! {donnees.get("count", 0)} enregistrement(s) trouvé(s).')
            
            return redirect('etats:detail', pk=etat.pk)
            
        except Exception as e:
            import traceback
            etat.statut = 'ERREUR'
            etat.save()
            
            # Logger l'erreur complète pour le débogage
            error_details = traceback.format_exc()
            print(f"Erreur lors de la génération de l'état {etat.pk}: {str(e)}")
            print(f"Détails complets: {error_details}")
            
            messages.error(request, f'Erreur lors de la génération: {str(e)}')
            return redirect('etats:detail', pk=etat.pk)
    
    def calculer_donnees(self, etat):
        """Calcule les données selon le type d'état"""
        donnees = {
            'total_usd': Decimal('0.00'),
            'total_cdf': Decimal('0.00'),
            'lignes': []
        }
        
        if etat.type_etat == 'DEMANDE_PAIEMENT':
            return self.calculer_donnees_demandes(etat)
        elif etat.type_etat == 'RECETTE':
            return self.calculer_donnees_recettes(etat)
        elif etat.type_etat == 'DEPENSE':
            return self.calculer_donnees_depenses(etat)
        elif etat.type_etat == 'PAIEMENT':
            return self.calculer_donnees_paiements(etat)
        elif etat.type_etat == 'RELEVE_DEPENSE':
            return self.calculer_donnees_releves_depenses(etat)
        elif etat.type_etat == 'SOLDE_BANCAIRE':
            return self.calculer_donnees_soldes_bancaires(etat)
        else:
            return donnees
    
    def calculer_donnees_demandes(self, etat):
        """Calcule les données pour les demandes de paiement"""
        queryset = DemandePaiement.objects.select_related(
            'service_demandeur', 'nature_economique', 'cree_par', 'approuve_par'
        )
        
        # Appliquer les filtres de période
        queryset = queryset.filter(
            date_soumission__date__gte=etat.date_debut,
            date_soumission__date__lte=etat.date_fin
        )
        
        # Appliquer les filtres supplémentaires
        filtres = etat.filtres_supplementaires
        if filtres.get('statut_demande'):
            queryset = queryset.filter(statut=filtres['statut_demande'])
        if filtres.get('devise'):
            queryset = queryset.filter(devise=filtres['devise'])
        if filtres.get('montant_min'):
            queryset = queryset.filter(montant__gte=filtres['montant_min'])
        if filtres.get('montant_max'):
            queryset = queryset.filter(montant__lte=filtres['montant_max'])
        
        # Appliquer les filtres many-to-many
        if etat.services.exists():
            queryset = queryset.filter(service_demandeur__in=etat.services.all())
        if etat.natures_economiques.exists():
            queryset = queryset.filter(nature_economique__in=etat.natures_economiques.all())
        
        # Appliquer le tri
        tri_par = etat.parametres_affichage.get('tri_par', 'date')
        ordre_tri = etat.parametres_affichage.get('ordre_tri', 'desc')
        
        if tri_par == 'date':
            queryset = queryset.order_by(f"-{'date_soumission' if ordre_tri == 'desc' else 'date_soumission'}")
        elif tri_par == 'montant':
            queryset = queryset.order_by(f"{'-' if ordre_tri == 'desc' else ''}montant")
        elif tri_par == 'reference':
            queryset = queryset.order_by(f"{'-' if ordre_tri == 'desc' else ''}reference")
        elif tri_par == 'service':
            queryset = queryset.order_by(f"{'-' if ordre_tri == 'desc' else ''}service_demandeur__nom_service")
        
        # Calculer les totaux
        total_usd = queryset.filter(devise='USD').aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
        total_cdf = queryset.filter(devise='CDF').aggregate(total=Sum('montant'))['total'] or Decimal('0.00')
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_recettes(self, etat):
        """Calcule les données pour les recettes"""
        queryset = Recette.objects.select_related('banque', 'compte_bancaire', 'enregistre_par')
        
        # Appliquer les filtres de période
        queryset = queryset.filter(
            date_encaissement__gte=etat.date_debut,
            date_encaissement__lte=etat.date_fin
        )
        
        # Appliquer les filtres supplémentaires
        filtres = etat.filtres_supplementaires
        if filtres.get('source_recette'):
            queryset = queryset.filter(source_recette=filtres['source_recette'])
        if filtres.get('montant_min'):
            queryset = queryset.filter(
                Q(montant_usd__gte=filtres['montant_min']) | 
                Q(montant_cdf__gte=filtres['montant_min'])
            )
        if filtres.get('montant_max'):
            queryset = queryset.filter(
                Q(montant_usd__lte=filtres['montant_max']) | 
                Q(montant_cdf__lte=filtres['montant_max'])
            )
        
        # Appliquer les filtres many-to-many
        if etat.banques.exists():
            queryset = queryset.filter(banque__in=etat.banques.all())
        if etat.comptes_bancaires.exists():
            queryset = queryset.filter(compte_bancaire__in=etat.comptes_bancaires.all())
        
        # Calculer les totaux
        total_usd = queryset.aggregate(total=Sum('montant_usd'))['total'] or Decimal('0.00')
        total_cdf = queryset.aggregate(total=Sum('montant_cdf'))['total'] or Decimal('0.00')
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_depenses(self, etat):
        """Calcule les données pour les dépenses"""
        queryset = Depense.objects.select_related('nomenclature', 'banque')
        
        # Appliquer les filtres de période
        queryset = queryset.filter(
            date_depense__gte=etat.date_debut,
            date_depense__lte=etat.date_fin
        )
        
        # Appliquer les filtres supplémentaires
        filtres = etat.filtres_supplementaires
        if filtres.get('code_depense'):
            queryset = queryset.filter(code_depense__icontains=filtres['code_depense'])
        
        # Appliquer les filtres many-to-many
        if etat.banques.exists():
            queryset = queryset.filter(banque__in=etat.banques.all())
        
        # Calculer les totaux
        total_usd = queryset.aggregate(total=Sum('montant_usd'))['total'] or Decimal('0.00')
        total_cdf = queryset.aggregate(total=Sum('montant_fc'))['total'] or Decimal('0.00')
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_paiements(self, etat):
        """Calcule les données pour les paiements"""
        queryset = Paiement.objects.select_related(
            'demande', 'releve_depense', 'paiement_par'
        )
        
        # Appliquer les filtres de période
        queryset = queryset.filter(
            date_paiement__date__gte=etat.date_debut,
            date_paiement__date__lte=etat.date_fin
        )
        
        # Appliquer les filtres supplémentaires
        filtres = etat.filtres_supplementaires
        if filtres.get('devise'):
            queryset = queryset.filter(devise=filtres['devise'])
        if filtres.get('montant_min'):
            queryset = queryset.filter(montant_paye__gte=filtres['montant_min'])
        if filtres.get('montant_max'):
            queryset = queryset.filter(montant_paye__lte=filtres['montant_max'])
        
        # Calculer les totaux
        total_usd = queryset.filter(devise='USD').aggregate(total=Sum('montant_paye'))['total'] or Decimal('0.00')
        total_cdf = queryset.filter(devise='CDF').aggregate(total=Sum('montant_paye'))['total'] or Decimal('0.00')
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_releves_depenses(self, etat):
        """Calcule les données pour les relevés de dépenses"""
        queryset = ReleveDepense.objects.select_related('valide_par').prefetch_related('demandes')
        
        # Debug: Afficher les informations de base
        print(f"=== DEBUG RELEVÉS ===")
        print(f"Période demandée: {etat.date_debut} au {etat.date_fin}")
        print(f"Type état: {etat.type_etat}")
        
        # Vérifier d'abord tous les relevés sans filtre
        tous_releves = ReleveDepense.objects.all()
        print(f"Total relevés dans la base: {tous_releves.count()}")
        
        # Afficher les périodes disponibles
        periodes_disponibles = tous_releves.values_list('periode', flat=True)
        print(f"Périodes disponibles: {list(periodes_disponibles)}")
        
        # CORRECTION: Pour les relevés de dépenses, la période est stockée comme premier jour du mois
        # Il faut donc inclure tous les relevés dont la période est entre date_debut et date_fin
        # en considérant que la période du relevé représente tout le mois
        
        # Si la période demandée est un seul jour, l'élargir au mois complet
        if etat.date_debut == etat.date_fin:
            # C'est une demande pour un jour spécifique, trouver le mois correspondant
            mois_debut = etat.date_debut.replace(day=1)
            # Dernier jour du mois
            if etat.date_debut.month == 12:
                mois_fin = etat.date_debut.replace(etat.date_debut.year + 1, 1, 1)
            else:
                mois_fin = etat.date_debut.replace(etat.date_debut.year, etat.date_debut.month + 1, 1)
            mois_fin = mois_fin - timezone.timedelta(days=1)
            
            print(f"Période élargie au mois: {mois_debut} au {mois_fin}")
            queryset = queryset.filter(
                periode__gte=mois_debut,
                periode__lte=mois_fin
            )
        else:
            # Période normale, mais s'assurer d'inclure les mois complets
            queryset = queryset.filter(
                periode__gte=etat.date_debut.replace(day=1),
                periode__lte=etat.date_fin
            )
        
        print(f"Relevés après filtre période: {queryset.count()}")
        
        # Le modèle ReleveDepense n'a pas de champ banque, donc pas de filtrage par banque
        # if etat.banques.exists():
        #     queryset = queryset.filter(banque__in=etat.banques.all())
        #     print(f"Relevés après filtre banques: {queryset.count()}")
        
        # Calculer les totaux
        total_usd = queryset.aggregate(total=Sum('net_a_payer_usd'))['total'] or Decimal('0.00')
        total_cdf = queryset.aggregate(total=Sum('net_a_payer_cdf'))['total'] or Decimal('0.00')
        
        print(f"Totaux calculés: USD={total_usd}, CDF={total_cdf}")
        print(f"========================")
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def calculer_donnees_soldes_bancaires(self, etat):
        """Calcule les données pour les soldes bancaires"""
        queryset = CompteBancaire.objects.select_related('banque').filter(actif=True)
        
        # Appliquer les filtres many-to-many
        if etat.banques.exists():
            queryset = queryset.filter(banque__in=etat.banques.all())
        if etat.comptes_bancaires.exists():
            queryset = queryset.filter(pk__in=etat.comptes_bancaires.all())
        
        # Calculer les totaux - le modèle CompteBancaire a un champ devise
        total_usd = queryset.filter(devise='USD').aggregate(total=Sum('solde_courant'))['total'] or Decimal('0.00')
        total_cdf = queryset.filter(devise='CDF').aggregate(total=Sum('solde_courant'))['total'] or Decimal('0.00')
        
        return {
            'total_usd': total_usd,
            'total_cdf': total_cdf,
            'lignes': queryset,
            'count': queryset.count()
        }
    
    def generer_pdf(self, etat, donnees):
        """Génère le fichier PDF pour l'état"""
        try:
            logger.info(f"=== DÉBUT GÉNÉRATION PDF ===")
            logger.info(f"État ID: {etat.pk}")
            logger.info(f"Type: {etat.type_etat}")
            logger.info(f"Nombre de lignes: {len(donnees.get('lignes', []))}")
            
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{etat.get_nom_fichier("pdf")}"'
            
            # Créer le document PDF sans pied de page automatique
            doc = SimpleDocTemplate(
                response, 
                pagesize=A4, 
                rightMargin=50,
                leftMargin=50,
                topMargin=50,
                bottomMargin=50
            )
            
            # Contenu du PDF
            story = []
            styles = getSampleStyleSheet()
            
            # Style pour le titre principal (type + période sur une ligne)
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=16,
                spaceAfter=20,
                alignment=TA_CENTER,
                textColor=colors.darkblue
            )
            
            # Titre principal
            titre_complet = self.generer_titre_complet(etat)
            story.append(Paragraph(titre_complet, title_style))
            
            story.append(Spacer(1, 20))
            
            # Tableau des données
            if donnees['lignes'] and donnees.get('count', 0) > 0:
                logger.info(f"Ajout du tableau pour {etat.type_etat}")
                self.ajouter_tableau_donnees(story, etat, donnees)
                logger.info("Tableau ajouté avec succès")
            else:
                # Message si aucune donnée
                story.append(Paragraph("AUCUNE DONNÉE TROUVÉE", styles['Heading3']))
                story.append(Paragraph(f"Aucun enregistrement trouvé pour la période du {etat.date_debut} au {etat.date_fin}.", styles['Normal']))
                story.append(Paragraph("Essayez d'élargir la période ou d'ajuster les filtres.", styles['Normal']))
                story.append(Spacer(1, 20))
            
            # Ajouter le pied de page manuellement
            story.append(Spacer(1, 30))
            
            # Ligne de séparation
            story.append(Spacer(1, 10))
            
            # Tableau pour le pied de page
            footer_data = [
                [f"Kinshasa, le {timezone.now().strftime('%d/%m/%Y')}", f"Page 1"]  # Page 1 pour l'instant
            ]
            
            footer_table = Table(footer_data, colWidths=[400, 100])
            footer_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 10),
                ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
                ('LINEABOVE', (0, 0), (-1, 0), 2, colors.black),
            ]))
            
            story.append(footer_table)
            
            # Construire le PDF
            logger.info("Construction du PDF...")
            doc.build(story)
            logger.info("PDF construit avec succès")
            
            # Sauvegarder le fichier
            from django.core.files.base import ContentFile
            etat.fichier_pdf.save(etat.get_nom_fichier('pdf'), ContentFile(response.content))
            etat.save()
            logger.info("Fichier PDF sauvegardé")
            logger.info(f"=== FIN GÉNÉRATION PDF ===")
            
        except Exception as e:
            logger.error(f"Erreur dans generer_pdf: {str(e)}")
            import traceback
            logger.error(traceback.format_exc())
            raise e
    
    def generer_titre_complet(self, etat):
        """Génère le titre complet avec format 'Liste des... du... au...'"""
        type_label = etat.get_type_etat_display()
        
        # Convertir le type en format "Liste des..." avec gestion correcte du pluriel
        type_lower = type_label.lower()
        if type_lower == 'paiement':
            titre_liste = "Liste des paiements"
        elif type_lower == 'demande':
            titre_liste = "Liste des demandes"
        elif type_lower == 'relevé':
            titre_liste = "Liste des relevés"
        elif type_lower == 'recette':
            titre_liste = "Liste des recettes"
        elif type_lower == 'dépense':
            titre_liste = "Liste des dépenses"
        elif type_lower == 'solde':
            titre_liste = "Liste des soldes"
        else:
            titre_liste = f"Liste des {type_label}"
        
        if etat.date_debut == etat.date_fin:
            # Même jour
            date_str = etat.date_debut.strftime('%d/%m/%Y')
            return f"{titre_liste} du {date_str}"
        else:
            # Période
            debut_str = etat.date_debut.strftime('%d/%m/%Y')
            fin_str = etat.date_fin.strftime('%d/%m/%Y')
            return f"{titre_liste} du {debut_str} au {fin_str}"
    
    def ajouter_tableau_donnees(self, story, etat, donnees):
        """Ajoute le tableau des données selon le type d'état"""
        if etat.type_etat == 'DEMANDE_PAIEMENT':
            self.ajouter_tableau_demandes(story, donnees['lignes'])
        elif etat.type_etat == 'RECETTE':
            self.ajouter_tableau_recettes(story, donnees['lignes'])
        elif etat.type_etat == 'DEPENSE':
            self.ajouter_tableau_depenses(story, donnees['lignes'])
        elif etat.type_etat == 'PAIEMENT':
            self.ajouter_tableau_paiements(story, donnees['lignes'])
        elif etat.type_etat == 'RELEVE_DEPENSE':
            self.ajouter_tableau_releves(story, donnees['lignes'])
        elif etat.type_etat == 'SOLDE_BANCAIRE':
            self.ajouter_tableau_soldes(story, donnees['lignes'])
    
    def ajouter_tableau_demandes(self, story, demandes):
        """Ajoute le tableau des demandes de paiement avec totaux"""
        headers = ['N°', 'Référence', 'Service', 'Nature', 'Montant', 'Devise', 'Statut']
        data = [headers]
        
        total_usd = Decimal('0.00')
        total_cdf = Decimal('0.00')
        
        for idx, demande in enumerate(demandes, 1):
            # Formatter le montant avec séparateurs
            montant_formate = f"{demande.montant or Decimal('0.00'):,.2f}".replace(',', ' ').replace('.', ',')
            
            # Formatter la nature économique: titre (code)
            nature_str = ""
            if demande.nature_economique:
                nature_str = f"{demande.nature_economique.titre}"
                if demande.nature_economique.code:
                    nature_str += f" ({demande.nature_economique.code})"
            
            data.append([
                str(idx),
                demande.reference or '',
                demande.service_demandeur.nom_service if demande.service_demandeur else '',
                nature_str,
                montant_formate,
                demande.devise or '',
                demande.get_statut_display() if hasattr(demande, 'get_statut_display') else ''
            ])
            
            # Calculer les totaux
            if demande.devise == 'USD':
                total_usd += demande.montant or Decimal('0.00')
            elif demande.devise == 'CDF':
                total_cdf += demande.montant or Decimal('0.00')
        
        # Ajouter la ligne de totaux
        data.append([
            'TOTAL',
            '',
            '',
            '',
            '',
            '',
            ''
        ])
        data.append([
            '',
            'Total USD',
            '',
            '',
            f"{total_usd:,.2f}".replace(',', ' ').replace('.', ','),
            'USD',
            ''
        ])
        data.append([
            '',
            'Total CDF',
            '',
            '',
            f"{total_cdf:,.2f}".replace(',', ' ').replace('.', ','),
            'CDF',
            ''
        ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Colonne N° centrée
            ('ALIGN', (4, 1), (4, -1), 'RIGHT'),  # Colonne montant à droite
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            # Style pour la ligne TOTAL
            ('BACKGROUND', (0, -3), (-1, -3), colors.darkblue),
            ('TEXTCOLOR', (0, -3), (-1, -3), colors.white),
            ('FONTNAME', (0, -3), (-1, -3), 'Helvetica-Bold'),
            # Style pour les lignes de totaux
            ('FONTNAME', (1, -2), (2, -1), 'Helvetica-Bold'),
        ]))
        
        story.append(table)
    
    def ajouter_tableau_recettes(self, story, recettes):
        """Ajoute le tableau des recettes avec totaux"""
        headers = ['N°', 'Référence', 'Banque', 'Compte', 'Source', 'Montant USD', 'Montant CDF', 'Date']
        data = [headers]
        
        total_usd = Decimal('0.00')
        total_cdf = Decimal('0.00')
        
        for idx, recette in enumerate(recettes, 1):
            montant_usd = Decimal(str(getattr(recette, 'montant_usd', '0')))
            montant_cdf = Decimal(str(getattr(recette, 'montant_cdf', '0')))
            
            # Formatter les montants avec séparateurs
            usd_formate = f"{montant_usd:,.2f}".replace(',', ' ').replace('.', ',')
            cdf_formate = f"{montant_cdf:,.2f}".replace(',', ' ').replace('.', ',')
            
            data.append([
                str(idx),
                getattr(recette, 'reference', '') or '',
                recette.banque.nom_banque if recette.banque else '',
                recette.compte_bancaire.intitule_compte if recette.compte_bancaire else '',
                getattr(recette, 'source', '') or '',
                usd_formate,
                cdf_formate,
                recette.date_encaissement.strftime('%d/%m/%Y') if hasattr(recette, 'date_encaissement') and recette.date_encaissement else ''
            ])
            
            # Calculer les totaux
            total_usd += montant_usd
            total_cdf += montant_cdf
        
        # Ajouter la ligne de totaux
        data.append([
            'TOTAL',
            '',
            '',
            '',
            '',
            '',
            '',
            ''
        ])
        data.append([
            '',
            'Total USD',
            '',
            '',
            '',
            f"{total_usd:,.2f}".replace(',', ' ').replace('.', ','),
            '',
            ''
        ])
        data.append([
            '',
            'Total CDF',
            '',
            '',
            '',
            '',
            f"{total_cdf:,.2f}".replace(',', ' ').replace('.', ','),
            ''
        ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Colonne N° centrée
            ('ALIGN', (5, 1), (6, -1), 'RIGHT'),  # Colonnes montants à droite
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            # Style pour la ligne TOTAL
            ('BACKGROUND', (0, -3), (-1, -3), colors.darkblue),
            ('TEXTCOLOR', (0, -3), (-1, -3), colors.white),
            ('FONTNAME', (0, -3), (-1, -3), 'Helvetica-Bold'),
            # Style pour les lignes de totaux
            ('FONTNAME', (1, -2), (2, -1), 'Helvetica-Bold'),
        ]))
        
        story.append(table)
    
    def ajouter_tableau_depenses(self, story, depenses):
        """Ajoute le tableau des dépenses avec totaux"""
        headers = ['Code', 'Libellé', 'Banque', 'Montant USD', 'Montant CDF', 'Date']
        data = [headers]
        
        total_usd = Decimal('0.00')
        total_cdf = Decimal('0.00')
        
        for depense in depenses:
            data.append([
                depense.code_depense or '',
                depense.libelle_depenses[:50] + '...' if len(depense.libelle_depenses) > 50 else depense.libelle_depenses,
                depense.banque.nom_banque if depense.banque else '',
                str(depense.montant_usd),
                str(depense.montant_fc),
                depense.date_depense.strftime('%d/%m/%Y') if depense.date_depense else ''
            ])
            
            # Calculer les totaux
            total_usd += depense.montant_usd
            total_cdf += depense.montant_fc
        
        # Ajouter la ligne de totaux
        data.append([
            'TOTAL',
            '',
            '',
            '',
            '',
            ''
        ])
        data.append([
            '',
            'Total USD',
            '',
            f"{total_usd:,.2f}",
            '',
            ''
        ])
        data.append([
            '',
            'Total CDF',
            '',
            '',
            f"{total_cdf:,.2f}",
            ''
        ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 1), (4, -1), 'RIGHT'),  # Colonnes montants à droite
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            # Style pour la ligne TOTAL
            ('BACKGROUND', (0, -3), (-1, -3), colors.darkblue),
            ('TEXTCOLOR', (0, -3), (-1, -3), colors.white),
            ('FONTNAME', (0, -3), (-1, -3), 'Helvetica-Bold'),
            # Style pour les lignes de totaux
            ('FONTNAME', (1, -2), (2, -1), 'Helvetica-Bold'),
        ]))
        
        story.append(table)
    
    def ajouter_tableau_paiements(self, story, paiements):
        """Ajoute le tableau des paiements avec totaux"""
        headers = ['N°', 'Référence', 'Demande', 'Montant', 'Devise', 'Date']
        data = [headers]
        
        total_usd = Decimal('0.00')
        total_cdf = Decimal('0.00')
        
        for idx, paiement in enumerate(paiements, 1):
            # Formatter la demande: DEM-000004 - Finance
            demande_str = ""
            if paiement.demande:
                demande_str = f"{paiement.demande.reference}"
                if paiement.demande.service_demandeur:
                    demande_str += f" - {paiement.demande.service_demandeur.nom_service}"
            
            # Formatter le montant avec séparateurs
            montant_formate = f"{paiement.montant_paye:,.2f}".replace(',', ' ').replace('.', ',')
            
            data.append([
                str(idx),
                paiement.reference or '',
                demande_str,
                montant_formate,
                paiement.devise or '',
                paiement.date_paiement.strftime('%d/%m/%Y') if paiement.date_paiement else ''
            ])
            
            # Calculer les totaux
            if paiement.devise == 'USD':
                total_usd += paiement.montant_paye
            elif paiement.devise == 'CDF':
                total_cdf += paiement.montant_paye
        
        # Ajouter la ligne de totaux
        data.append([
            'TOTAL',
            '',
            '',
            '',
            '',
            ''
        ])
        data.append([
            '',
            'Total USD',
            f"{total_usd:,.2f}".replace(',', ' ').replace('.', ','),
            'USD',
            '',
            ''
        ])
        data.append([
            '',
            'Total CDF',
            f"{total_cdf:,.2f}".replace(',', ' ').replace('.', ','),
            'CDF',
            '',
            ''
        ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # Colonne N° centrée
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),  # Colonne montant à droite
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            # Style pour la ligne TOTAL
            ('BACKGROUND', (0, -3), (-1, -3), colors.darkblue),
            ('TEXTCOLOR', (0, -3), (-1, -3), colors.white),
            ('FONTNAME', (0, -3), (-1, -3), 'Helvetica-Bold'),
            # Style pour les lignes de totaux
            ('FONTNAME', (1, -2), (2, -1), 'Helvetica-Bold'),
        ]))
        
        story.append(table)
    
    def ajouter_tableau_soldes(self, story, comptes):
        """Ajoute le tableau des soldes bancaires"""
        headers = ['Banque', 'Compte', 'Devise', 'Solde courant']
        data = [headers]
        
        for compte in comptes:
            data.append([
                compte.banque.nom_banque if compte.banque else '',
                compte.intitule_compte or '',
                compte.devise or '',
                str(compte.solde_courant)
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),  # Colonne solde à droite
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
        ]))
        
        story.append(table)
    
    def ajouter_tableau_releves(self, story, releves):
        """Ajoute le tableau des relevés de dépenses avec totaux"""
        headers = ['Numéro', 'Période', 'Net USD', 'Net CDF', 'Date création']
        data = [headers]
        
        total_usd = Decimal('0.00')
        total_cdf = Decimal('0.00')
        
        for releve in releves:
            data.append([
                releve.numero or '',
                str(releve.periode) if releve.periode else '',
                str(releve.net_a_payer_usd),
                str(releve.net_a_payer_cdf),
                releve.date_creation.strftime('%d/%m/%Y') if releve.date_creation else ''
            ])
            
            # Calculer les totaux
            total_usd += releve.net_a_payer_usd
            total_cdf += releve.net_a_payer_cdf
        
        # Ajouter la ligne de totaux
        data.append([
            'TOTAL',
            '',
            '',
            '',
            ''
        ])
        data.append([
            '',
            'Total USD',
            f"{total_usd:,.2f}",
            '',
            ''
        ])
        data.append([
            '',
            'Total CDF',
            '',
            f"{total_cdf:,.2f}",
            ''
        ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (2, 1), (3, -1), 'RIGHT'),  # Colonnes montants USD/CDF à droite
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
            # Style pour la ligne TOTAL
            ('BACKGROUND', (0, -3), (-1, -3), colors.darkblue),
            ('TEXTCOLOR', (0, -3), (-1, -3), colors.white),
            ('FONTNAME', (0, -3), (-1, -3), 'Helvetica-Bold'),
            # Style pour les lignes de totaux
            ('FONTNAME', (1, -2), (2, -1), 'Helvetica-Bold'),
        ]))
        
        story.append(table)
    
    def generer_excel(self, etat, donnees):
        """Génère le fichier Excel pour l'état"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
        
        wb = Workbook()
        ws = wb.active
        ws.title = etat.get_type_etat_display()
        
        # En-tête
        ws['A1'] = 'DGRAD - DIRECTION GÉNÉRALE DES REVENUS'
        ws['A2'] = etat.get_type_etat_display()
        ws['A3'] = f'Titre: {etat.titre}'
        ws['A4'] = f'Période: {etat.date_debut} au {etat.date_fin}'
        
        # Styles
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1a3a5f', end_color='1a3a5f', fill_type='solid')
        
        # Ajouter les données selon le type d'état
        row = 8
        if etat.type_etat == 'DEMANDE_PAIEMENT':
            headers = ['Référence', 'Service', 'Nature', 'Description', 'Montant', 'Devise', 'Statut']
            ws.append(headers)
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
            
            for demande in donnees['lignes']:
                row += 1
                ws.append([
                    demande.reference,
                    demande.service_demandeur.nom_service if demande.service_demandeur else '',
                    demande.nature_economique.code if demande.nature_economique else '',
                    demande.description,
                    float(demande.montant),
                    demande.devise,
                    demande.get_statut_display()
                ])
        
        # Sauvegarder le fichier
        from io import BytesIO
        output = BytesIO()
        wb.save(output)
        
        from django.core.files.base import ContentFile
        etat.fichier_excel.save(etat.get_nom_fichier('xlsx'), ContentFile(output.getvalue()))
        etat.save()


class EtatTelechargerView(LoginRequiredMixin, View):
    """Vue pour télécharger un état généré"""
    
    def get(self, request, pk, format_file):
        etat = get_object_or_404(EtatGenerique, pk=pk)
        
        if format_file == 'pdf' and etat.fichier_pdf:
            response = HttpResponse(etat.fichier_pdf.read(), content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{etat.get_nom_fichier("pdf")}"'
            return response
        elif format_file == 'excel' and etat.fichier_excel:
            response = HttpResponse(etat.fichier_excel.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="{etat.get_nom_fichier("xlsx")}"'
            return response
        else:
            messages.error(request, 'Fichier non disponible')
            return redirect('etats:detail', pk=etat.pk)
